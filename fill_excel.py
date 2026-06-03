"""fill_excel.py v4 — 보장분석 JSON → 색칠 보장진단 엑셀 (법칙 33조 반영).
- 계약 최대 50건: 7칸 초과분 합계열 앞에 자동 삽입.
- 헤더 갱신=파랑배경+흰글자 / 비갱신=빨강배경+흰글자.
- 값셀 갱신=파랑글자 / 비갱신=검정글자.
- 폼(B열)에 없는 담보·중복담보(예: 암진단비/암진단비(갱신))는 하단에 '➕추가담보' 새 행으로 추가.
- 메모(셀프팩폭·증권확인)는 '📋확인사항' 별도 시트에 분류색으로.
"""
import re, openpyxl
from copy import copy
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE="FF0000FF"; RED="FFFF0000"; BLACK="FF000000"; WHITE="FFFFFFFF"
F_BLUE=PatternFill("solid", fgColor=BLUE); F_RED=PatternFill("solid", fgColor=RED)
FIRST=3; TPL_COLS=7; MAXC=50
CIRC="①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳"
THIN=Side(style="thin", color="FFBFBFBF")
BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

def _norm(s):
    s=re.sub(r"\s+","",str(s or ""))
    return re.sub(r"[()（）·.\-_/]","",s)

def _renew_contract(c):
    rv=str(c.get("renewal",""))
    return ("비" not in rv) and ("갱신" in rv)

def _mark(i): return CIRC[i] if i<len(CIRC) else f"{i+1}."

def fill_excel(data, template_path, out_path):
    wb=openpyxl.load_workbook(template_path); ws=wb.active
    client=data.get("client","고객")
    _safe=re.sub(r"[:\\/?*\[\]]","",str(client)).strip() or "고객"
    ws.title=(_safe+" 보장진단")[:31]
    ws["A1"]=f"{client}\n보장진단"
    contracts=data.get("contracts",[])[:MAXC]
    N=len(contracts)

    extra=max(0, N-TPL_COLS)
    if extra:
        ws.insert_cols(10, extra)
        for newc in range(10, 10+extra):
            for r in range(1, ws.max_row+1):
                dst=ws.cell(r,newc)
                if not isinstance(dst,MergedCell):
                    dst._style=copy(ws.cell(r,FIRST)._style)
        w=ws.column_dimensions[get_column_letter(FIRST)].width or 9
        for newc in range(10,10+extra):
            ws.column_dimensions[get_column_letter(newc)].width=w
    last_col=FIRST+N-1 if N else FIRST
    sum_col=10+extra; memo_col=sum_col+1

    for i,c in enumerate(contracts):
        col=FIRST+i; ren=_renew_contract(c)
        h=ws.cell(1,col)
        h.value=f"{_mark(i)}{c.get('company','')}\n{c.get('product','')}"
        h.fill=F_BLUE if ren else F_RED
        h.font=Font(bold=True, color=WHITE, size=9)
        h.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(2,col).value=c.get("renewal","")
        ws.cell(3,col).value=c.get("premium")
        ws.cell(4,col).value=c.get("start") or _split(c.get("period"),0)
        ws.cell(5,col).value=c.get("end")   or _split(c.get("period"),1)
        ws.cell(6,col).value=c.get("payTerm")
        ws.cell(7,col).value=c.get("payCount")

    def setsum(r):
        cell=ws.cell(r,sum_col)
        if not isinstance(cell,MergedCell) and N:
            cell.value=f"=SUM({get_column_letter(FIRST)}{r}:{get_column_letter(last_col)}{r})"
    ws.cell(1,sum_col).value="합계"
    setsum(3)

    # 담보명 → 행
    name2row={}
    for r in range(8, ws.max_row+1):
        b=ws.cell(r,2).value
        if b: name2row.setdefault(_norm(b), r)

    filled=set()       # (row,col) 이미 채운 칸
    extra_rows=[]      # 폼에 없거나 중복인 담보 → 하단 추가 (group,name,amounts)
    for rw in data.get("coverage",[]):
        key=_norm(rw.get("name"))
        row=name2row.get(key) or next((v for k,v in name2row.items() if key and (key in k or k in key)), None)
        collide=False
        amts=rw.get("amounts",[])
        if row:
            for a in amts:
                ci=a.get("ci")
                if ci is None or ci>=N: continue
                cell=ws.cell(row, FIRST+ci)
                if isinstance(cell,MergedCell): continue
                if (row,FIRST+ci) in filled:   # 중복담보 → 추가행으로
                    collide=True; break
            if not collide:
                for a in amts:
                    ci=a.get("ci")
                    if ci is None or ci>=N: continue
                    cell=ws.cell(row, FIRST+ci)
                    if isinstance(cell,MergedCell): continue
                    cell.value=a.get("value")
                    cell.font=Font(color=BLUE if a.get("renew") else BLACK)
                    filled.add((row,FIRST+ci))
        if (not row) or collide:
            extra_rows.append((rw.get("group",""), rw.get("name",""), amts))

    # 합계열 공식(템플릿 담보행)
    for r in range(8, ws.max_row+1):
        if ws.cell(r,2).value: setsum(r)

    # ➕ 추가담보(폼 외/중복) 하단에 행 추가
    if extra_rows:
        sr=ws.max_row+2
        hdr=ws.cell(sr,1,"➕추가담보(폼 외/중복)")
        hdr.font=Font(bold=True, color="FF7F6000"); hdr.fill=PatternFill("solid",fgColor="FFFFF2CC")
        ws.merge_cells(start_row=sr,start_column=1,end_row=sr,end_column=2)
        sr+=1
        for grp,name,amts in extra_rows:
            ws.cell(sr,1,grp).font=Font(bold=True, size=9, color="FF595959")
            bcell=ws.cell(sr,2,name); bcell.font=Font(bold=True); bcell.border=BORDER
            ws.cell(sr,1).border=BORDER
            for a in amts:
                ci=a.get("ci")
                if ci is None or ci>=N: continue
                cell=ws.cell(sr, FIRST+ci)
                cell.value=a.get("value")
                cell.font=Font(color=BLUE if a.get("renew") else BLACK)
                cell.border=BORDER
            sc=ws.cell(sr,sum_col,f"=SUM({get_column_letter(FIRST)}{sr}:{get_column_letter(last_col)}{sr})")
            sc.border=BORDER
            sr+=1

    anchor=ws.cell(8,memo_col)
    if not isinstance(anchor,MergedCell):
        anchor.value="확인사항·메모는\n'📋확인사항' 시트 참조"
        anchor.alignment=Alignment(wrap_text=True, vertical="top")
        anchor.font=Font(size=9, italic=True, color="FF808080")

    memos=list(data.get("memo",[]))
    if N>=MAXC: memos.append(f"계약이 {MAXC}건을 초과해 앞 {MAXC}건만 표시함(증권 확인).")
    if extra_rows: memos.append(f"폼에 없던/중복 담보 {len(extra_rows)}건을 하단 '➕추가담보'에 추가함.")
    _memo_sheet(wb, client, memos)

    wb.save(out_path); return out_path

def _tagcolor(tag):
    t=str(tag)
    if "확인" in t: return ("⚠확인","FFFFD966","FF7F6000")
    if "부족" in t: return ("부족","FFF4B183","FF833C00")
    if "미가입" in t: return ("미가입","FFD9D9D9","FF595959")
    if "비갱신" in t: return ("비갱신","FFFCE4D6","FFC00000")
    if "갱신" in t: return ("갱신","FFBDD7EE","FF1F4E79")
    return ("메모","FFFFFFFF","FF000000")
def _memo_parts(m):
    # 신형 {tag,item,note} / 구형 문자열 모두 처리
    if isinstance(m, dict):
        return m.get("tag","메모"), str(m.get("item","")), str(m.get("note",""))
    t=str(m); tag="메모"
    for k in ("약관확인","확인 요망","확인요망","확인 필요","확인필요"):
        if k in t: tag="확인"; break
    else:
        if "부족" in t: tag="부족"
        elif "미가입" in t: tag="미가입"
        elif "비갱신" in t: tag="비갱신"
        elif "갱신" in t: tag="갱신"
    return tag, "", t

def _memo_sheet(wb, client, memos):
    ws=wb.create_sheet("📋확인사항")
    ws.column_dimensions["A"].width=11; ws.column_dimensions["B"].width=24; ws.column_dimensions["C"].width=44
    ws["A1"]=f"{client} · 셀프팩폭 / 증권확인"
    ws["A1"].font=Font(bold=True, size=13, color="FFFFFFFF")
    ws["A1"].fill=PatternFill("solid", fgColor="FF1F1F1F")
    ws.merge_cells("A1:C1"); ws.row_dimensions[1].height=26
    ws["A1"].alignment=Alignment(vertical="center", horizontal="left", indent=1)
    for c,t in (("A2","분류"),("B2","대상"),("C2","한 줄")):
        ws[c]=t; ws[c].font=Font(bold=True, color="FFFFFFFF"); ws[c].fill=PatternFill("solid",fgColor="FF595959")
        ws[c].alignment=Alignment(horizontal="center", vertical="center"); ws[c].border=BORDER
    r=3
    for m in memos:
        tag,item,note=_memo_parts(m); lab,bg,fg=_tagcolor(tag)
        a=ws.cell(r,1,lab); b=ws.cell(r,2,item); c=ws.cell(r,3,note)
        a.fill=PatternFill("solid", fgColor=bg)
        a.font=Font(bold=True, color=fg, size=10); b.font=Font(bold=True, color="FF222222", size=10); c.font=Font(color="FF222222", size=10)
        a.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        b.alignment=Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        c.alignment=Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        for cc in (a,b,c): cc.border=BORDER
        ws.row_dimensions[r].height=max(20, 15*(1+max(len(item),len(note))//26)); r+=1
    ws.freeze_panes="A3"; return ws

def _split(period,idx):
    if not period: return None
    m=re.split(r"[~\-]", str(period).split("/")[0])
    return m[idx].strip() if idx<len(m) else None
