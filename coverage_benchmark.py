# ===== BARUM coverage_benchmark.py v474-prodname-20260818 (구 v33-ci-rate-20260708 계승) =====
# -*- coding: utf-8 -*-
"""
BARUM 충족률 엔진 + map_excel_to_report
- 입력: 완성된 보장진단 엑셀(.xlsx) 1개 (등식2: 리포트는 완성 엑셀만 읽음)
- 출력: report_weasy.build_report_pdf 가 먹는 rep dict
- 충족률 % = 보유합 / 연령밴드별 권장액 * 100 (상한 없음·실제치, 2026.07.12 지점장 확정)
나이/성별 미추출 시 기본밴드 '40s' 적용 + gap_count/배지에 [확인] 노출.
"""
import re, openpyxl

# ── 충족률 권장 벤치마크 (단위: 만원). 출처 주석.
#   진단계열만 금액기반, 비금액계열(운전자/실손/일당/응급)은 presence 가중.
# SOURCES:
#   암  : 일반암 적정 5,000 + 고액암(중위소득×3, 최대 1억~1.5억) → 권장 합 밴드별
#         (banksalad 암보험 진단금 가이드)
#   뇌혈관: 최소 3,000, 40대 이후 4,000~5,000 (cancerok/뇌혈관 적정수준)
#   심장  : 허혈성 평균 설계 3,000 (banksalad 심장질환 11개사 비교)
#   수술비: 중수술 500~1,000 + 종수술 (signalplanner 리모델링 표)
#   사망·후유: 가장세대 보장공백 대비, 밴드별 상향
BENCHMARK = {
    #          사망후유  암    뇌혈관  심장  수술비  사망후유는 '사망후유' 키
    '20s': {'사망·후유':5000, '암':8000,  '뇌혈관':3000, '심장':3000},
    '30s': {'사망·후유':10000,'암':10000, '뇌혈관':4000, '심장':3000},
    '40s': {'사망·후유':15000,'암':10000, '뇌혈관':5000, '심장':3000},
    '50s': {'사망·후유':10000,'암':10000, '뇌혈관':5000, '심장':4000},
    '60s': {'사망·후유':5000, '암':8000,  '뇌혈관':5000, '심장':4000},
}
# ★★★★★v418 담보명 기준 벤치마크 (지점장 지시 2026.08.13 「담보이름으로」)
#   충족률의 「보유」 = <b>A열 그룹 세로합이 아니라 B열 담보명 하나의 값</b>.
#   출처: 뇌·심 4종 = <b>바름 교육자료(26년 바름 교육 2607) 「고객 진단비/치료비 보장 점검」</b> 권장액 실측.
#         암·사망 = 롯데 리포트 「한장보장현황」 표준금액 + 지점장 지시 「기준점을 높여라」.
#   ★담보가 2개인 카테고리는 <b>낮은 쪽</b>을 본다 — 충족률은 「어디가 비었나」를 보는 지표다.
NAMED_BENCH = {
    '뇌혈관':    [('뇌혈관진단비', 5000)],                      # 바름 교육 5,000
    '심장':      [('허혈성 진단비', 5000), ('급성심근경색', 5000)],  # 바름 교육 각 5,000
    '암':        [('일반암', 5000)],                            # 바름 교육 「암 진단비 5천~1억」 하단
    '사망·후유': [('상해사망', 20000)],                          # 롯데 표준 20,000
}
# 비금액(presence) 카테고리: 핵심담보 보유개수 / 기준개수
PRESENCE = {
    '운전자':    {'keys':['합의금','변호사','대인','대물','자부상','6주미만'], 'need':4},
    '입원·일당': {'keys':['간병인','상해일당','질병일당','간호통합','중환자'], 'need':3},
    '실손·일배책':{'keys':['입원','통원','약값','일상배상','일배책','MRI','상해의료비'],'need':3},   # ★v86 상해의료비 = 마스터 99행 신설, 보장현황·보장진단서에 그대로 표기   # ★v30 행명 '입원' 매칭
    '골절·화상': {'keys':['골절','화상','깁스','5대골절','중증화상'],          'need':3},
    # ★v30i 수술비 = 종·핵심수술담보 '개수' 기반(금액합·종슬래시 최댓값 폐기). 몇 종·몇 개 가입했는지가 기준.
    '수술비':    {'keys':['수술','창상'],                                    'need':4},
    '응급실·독감':{'keys':['응급실','독감','식중독','骨'],                      'need':2},
}
# 리포트 10카테고리 ← 엑셀 A열 15그룹
CATEGORY_GROUPS = {
    '사망·후유':  ['사망','후유장애'],
    '암':        ['암'],
    '뇌혈관':     ['뇌혈관'],
    '심장':       ['심장'],
    '수술비':     ['수술비','수술'],
    '운전자':     ['운전자','운전'],
    '입원·일당':  ['일당'],
    '실손·일배책':['실손','일배책'],
    '골절·화상':  ['골 절','골절','화상','깁스'],
    '응급실·독감':['응급실','독감'],
}
DONUT_ORDER = ['암','운전자','실손·일배책','수술비','뇌혈관','사망·후유','골절·화상','심장','입원·일당','응급실·독감']

def _man(v):
    try: return float(v)
    except:
        # ★v30 슬래시 행(종수술 1-5/1-8종·MRI트리오): 대표값 = 최대 칸
        try:
            if isinstance(v,str) and '/' in v:
                return max(float(p) for p in v.split('/') if p.strip().replace('.','').isdigit())
        except: pass
        # ★★v199: 완납 계약의 보험료 칸은 '58,340 (완납)' 텍스트다(엑셀 =SUM에서 자동 제외용).
        #   설명서 계약별 보험료가 0원으로 표기되던 것 차단 — 숫자만 뽑아 되돌린다.
        try:
            if isinstance(v,str):
                _m = re.match(r'^\s*([\d,]+)', v)
                if _m: return float(_m.group(1).replace(',',''))
        except: pass
        return 0.0

def _fmt(man):
    """만원 정수 → '1억7,100만' 한글단위"""
    man=int(round(man))
    if man<=0: return ''
    eok, rest = divmod(man,10000)
    s=''
    if eok: s+=f'{eok}억'
    if rest: s+=f'{rest:,}만'
    elif not eok: s='0'
    return s

def _bundle_adjust(path):
    """★v30e: [근거] 심장 섹션에서 묶음(여러 행 전개) 담보의 (행수-1)×금액을 카테고리별로 산출 → 보유합 중복 차감."""
    adj={'심장':0.0,'뇌혈관':0.0}
    try:
        wb=openpyxl.load_workbook(path,data_only=True)
        _qn=next((n for n in ('확인사항','📋확인사항') if n in wb.sheetnames), None)   # ★v41 시트명 이모지 제거 호환
        if not _qn: return adj
        qs=wb[_qn]; on=False
        for r in range(1,qs.max_row+1):
            a=str(qs.cell(r,1).value or '')
            if '[근거] 심장' in a: on=True; continue
            if on:
                rows=str(qs.cell(r,3).value or ''); amt=qs.cell(r,4).value
                if not rows or rows=='기재 행': continue
                if '·' in rows and isinstance(amt,(int,float)):
                    parts=[p.strip() for p in rows.split('·')]
                    n_h=sum(1 for p in parts if p not in ('뇌졸증진단비','뇌혈관진단비'))
                    n_b=len(parts)-n_h
                    if n_h>1: adj['심장']+=amt*(n_h-1)
                    if n_b>1: adj['뇌혈관']+=amt*(n_b-1)
    except Exception: pass
    return adj


def _CI_MAX(): return 3   # ★★★v236 영구지침(지점장 확정 2026.07.25): CI 계약 표기 <b>최대 3건</b>

def _ci_meta_list(path):
    """★★★v236 (지점장 확정 2026.07.25, 영구): <b>CI 계약을 계약별로 각각 표기한다 — 최대 3건</b>.
    구 `_ci_meta`는 `c=cols[0]`로 <b>첫 CI 열만</b> 봐서 두 번째 CI 계약을 통째로 버렸다
    (한정환 실측: 신한 라이프케어CI + DB CI종신 2건 중 DB가 사라지고 3p가 '선지급 80%형 사망 2,000만'으로 오출력).
    ★사망액 찾기: CI 종신은 담보명이 '일반사망'이 아니라 <b>'질병사망'</b>으로 인쇄되는 경우가 있어
      마스터 <b>일반사망 · 질병사망(80세)</b> 두 행을 모두 보고 <b>큰 값</b>을 사망액으로 쓴다.
      (구 코드는 사망액을 `bonche+중대한CI적용`으로 <b>역산</b>해 실제 10,000/5,000을 못 봤다.)
    ★선지급률: 정본은 <b>50% 또는 80% 두 가지뿐 · 추측 금지</b>.
      → 비율이 둘 중 어디에도 근접하지 않거나(±12% 초과) <b>본체 ≥ 사망액</b>이면
      선지급형이 아니라고 보고 <b>pct=None([확인])</b>으로 둔다. 억지 반올림 금지.
    """
    wb=openpyxl.load_workbook(path,data_only=True); ws=wb.active
    last=ws.max_column
    cols=[c for c in range(3,last) if _isci_hdr(ws.cell(1,c).value)]
    if not cols: return []
    rows={}
    for r in range(6,ws.max_row+1):
        b=str(ws.cell(r,2).value or '').strip()
        # ★★★v243(지점장 지시): CI 뇌 축은 <b>뇌졸증 or 뇌출혈</b> — '중대한 뇌출혈'도 반드시 본다.
        if b in ('중대한 암','중대한 뇌졸증','중대한 뇌출혈','중대한 급성심근','중대한CI적용',
                 '일반사망','질병사망(80세)','상해사망'): rows[b]=r
    def _cv(k,c):
        return _man(ws.cell(rows[k],c).value) if k in rows else 0
    out=[]
    for c in cols[:_CI_MAX()]:
        hdr=[x.strip() for x in str(ws.cell(1,c).value or '').split('\n') if x.strip()]
        company = hdr[0] if hdr else ''
        product = hdr[1] if len(hdr)>1 else ''
        # ★★★v243: <b>일반사망(종신)=CI 주계약</b>이 있으면 그것이 사망보장이다.
        #   구 max()는 신한(일반사망 4,000 / 질병사망 6,000=추가특약)에서 <b>6,000</b>을 골라
        #   2,000/6,000=33% → [확인]이 됐다. 주계약 4,000이면 50%형이 정확히 맞는다.
        samang = _cv('일반사망',c) or _cv('질병사망(80세)',c)
        bonche = max(_cv('중대한 암',c), _cv('중대한 뇌졸증',c), _cv('중대한 뇌출혈',c))
        resid  = _cv('중대한CI적용',c)
        pct=None; raw=0
        if samang>0 and bonche>0 and bonche<samang:
            raw=bonche/samang*100
            if min(abs(raw-80),abs(raw-50))<=12:      # ★50/80 근접만 인정
                pct = 80 if abs(raw-80)<=abs(raw-50) else 50
        out.append({'company':company,'product':product,
                    'samang':samang,'bonche':bonche,'resid':resid,
                    'pct':pct,'raw':round(raw),
                    'items':[(n,_cv(n,c)) for n in ('중대한 암','중대한 뇌졸증','중대한 뇌출혈','중대한 급성심근') if _cv(n,c)>0]})
    return out


def _isci_hdr(t):
    """★v235/v236 CI 상품명 판정 — main.py `_isci_prod`와 정본 1개(연속문자열 매칭 금지)."""
    _t0=str(t or ''); t=re.sub(r'[\s\u3000]','',_t0)
    if ('퍼펙트' in t) or ('퍼텍트' in t) or ('리빙케어' in t): return True
    # ★v246: 공백 제거 시 회사 영문약자와 CI가 붙는다(`무배당 KB CI종신보험`→`무배당KBCI종신보험`)
    #   → 영문 경계에 걸려 False가 됐다. <b>공백제거본과 원문 둘 다</b> 검사한다.
    _p=r'(?<![A-Za-z])(CI|GI)(?![A-Za-z])'
    return bool(re.search(_p, t) or re.search(_p, _t0)) and ('보험' in t or '종신' in t)


def _ci_meta(path):
    """★v33 선지급률 정본 계산 — CI 계약 '열'에서 직접 읽는다.
    끝열 '중대한CI적용' 은 비CI 계약 일반사망이 합산돼 오염되어 있어 사용 금지.
    지침: 선지급률은 50% 또는 80% 두 가지뿐."""
    wb=openpyxl.load_workbook(path,data_only=True); ws=wb.active
    last=ws.max_column
    def _isci(t):
        # ★★★v149 (지점장 지적 2026.07.21): CI 판정이 main.py `_isci_prod`에만 있고
        #   coverage_benchmark에는 <b>퍼펙트 예외가 통째로 빠져</b> 있었다 → 엑셀·PPT는 CI로 인식하는데
        #   보장진단서(설명서 3p 핵심보장)만 'CI보험 아님'으로 떴다. 두 경로의 판정을 일치시킨다.
        #   영구지침(2026.07.20): 삼성생명 <퍼펙트플러스보험>·<퍼펙트통합보험>은 표기 없어도 무조건 CI.
        t=re.sub(r'[\s\u3000]','',str(t or ''))
        # ★v150 '퍼펙트' 시리즈 전체 CI(퍼펙트통합·퍼펙트플러스·퍼펙트플러스종합 등 변형 포함)
        # ★★★v235 (한정환 실측): 구 `'CI보험'` 연속매칭이 `CI종신보험`을 놓쳤다(CI와 '보험' 사이에
        #   '종신' 등 수식어가 끼면 탈락). main.py `_isci_prod`와 <b>정본 1개로 통일</b>.
        #   영문 경계 필수 — `ACCIDENT` 안의 우연한 'CI' 배제.
        if ('퍼펙트' in t) or ('퍼텍트' in t) or ('리빙케어' in t): return True
        return bool(re.search(r'(?<![A-Za-z])(CI|GI)(?![A-Za-z])', t)) and ('보험' in t or '종신' in t)
    cols=[c for c in range(3,last) if _isci(ws.cell(1,c).value)]
    if not cols: return None
    rows={}
    for r in range(6,ws.max_row+1):
        b=str(ws.cell(r,2).value or '').strip()
        if b in ('중대한 암','중대한 뇌졸증','중대한 뇌출혈','중대한 급성심근','중대한CI적용'): rows[b]=r
    c=cols[0]
    # 본체 = 중대한 암·뇌졸증 (급성심근은 CI추가보장특약이 가산돼 오염)
    pure=[_man(ws.cell(rows[k],c).value) for k in ('중대한 암','중대한 뇌졸증','중대한 뇌출혈') if k in rows]
    pure=[v for v in pure if v>0]
    if not pure: return None
    bonche=max(pure)
    resid=_man(ws.cell(rows['중대한CI적용'],c).value) if '중대한CI적용' in rows else 0
    samang=bonche+resid
    if not samang: return None
    raw=bonche/samang*100
    pct=80 if abs(raw-80)<=abs(raw-50) else 50     # ★50/80 두 가지뿐
    return {'bonche':bonche,'samang':samang,'resid':resid,'pct':pct,'raw':round(raw)}


def load_excel(path):
    """완성 엑셀 → (groups_rows, headers). groups_rows[cat]=[(담보명,끝열값),..]"""
    wb=openpyxl.load_workbook(path,data_only=True); ws=wb.active
    last=ws.max_column
    # 그룹 경계: A열에 구분명 등장하는 행
    bounds=[]
    for r in range(6,ws.max_row+1):
        a=ws.cell(r,1).value
        if a and str(a).strip(): bounds.append((r,str(a).strip()))
    bounds.append((ws.max_row+1,'__END__'))
    grp_rows={}  # 엑셀그룹명 → [(담보,값)]
    disp={}      # ★v30h 담보명 → 슬래시 원문 표시(진단서에서 합산·최댓값 대신 가로 그대로)
    for i in range(len(bounds)-1):
        r0,name=bounds[i]; r1=bounds[i+1][0]
        rows=[]
        for r in range(r0,r1):
            b=ws.cell(r,2).value; _raw=ws.cell(r,last).value; v=_man(_raw)
            if b:
                _b=str(b).strip(); rows.append((_b,v))
                if isinstance(_raw,str) and '/' in _raw and any(c.isdigit() for c in _raw):
                    disp[_b]='('+_raw.strip()+')'   # 종수술비(1-5종)·n대·MRI = 슬래시 원문
        grp_rows[name]=rows
    # ★v30e 등식3: 혈전용해치료비(심장, 일당 그룹 첫 행 구조) → 심장으로 재배치 / 뇌쪽 혈전용해는 뇌혈관 그룹에 이미 소속
    if '일당' in grp_rows:
        _mv=[(b,v) for b,v in grp_rows['일당'] if '혈전용해' in b]
        if _mv:
            grp_rows['일당']=[(b,v) for b,v in grp_rows['일당'] if '혈전용해' not in b]
            grp_rows.setdefault('심장',[]).extend(_mv)
    load_excel._disp=disp   # ★v30h caller에 표시맵 전달(반환 시그니처 불변 → 회귀 0)
    # ★2026.07.11 실손 세대 자동판별용: '실손' 구분 그룹의 코어 담보(입원·통원·약값·의료비) 행
    # ★★★v250 (지점장 확정 2026.07.26): 3대비급여(도수·체외충격파·증식 / 비급여주사 / 비급여MRI)
    #   특약이 <b>분리돼 있으면 실손 세대 하한 = 3세대</b>. 이 3특약은 2017.04 3세대(착한실손)부터
    #   신설됐으므로 2세대 이하 계약에는 구조적으로 없다. main.py `_has_nonpay3`와 같은 규칙.
    _np3_rows=[r for r in range(6, ws.max_row+1)
               if ws.cell(r,2).value and (('도수' in str(ws.cell(r,2).value))
                                          or ('MRI' in str(ws.cell(r,2).value).upper())
                                          or ('체외충격파' in str(ws.cell(r,2).value))
                                          or ('증식' in str(ws.cell(r,2).value)))]
    load_excel._np3_rows=_np3_rows
    # ★v381 처방조제료(약값) 행 — 지점장 확정 "4세대는 통원비 20만원만 있다".
    #   약값이 통원과 별도로 잡혀 있으면 4세대가 아니다(3세대).
    load_excel._drug_rows=[r for r in range(6, ws.max_row+1)
                           if str(ws.cell(r,2).value or '').strip() in ('약값','처방조제료')]
    _sil_rows=[]
    for _i in range(len(bounds)-1):
        _r0,_nm=bounds[_i]; _r1=bounds[_i+1][0]
        if '실손' in str(_nm):
            _sil_rows=[r for r in range(_r0,_r1)
                       if ws.cell(r,2).value
                       and any(k in str(ws.cell(r,2).value) for k in ('입원','통원','약값','의료비'))
                       and not any(x in str(ws.cell(r,2).value) for x in ('MRI','도수','비급여주사','일당'))]
            break
    # ★v80 실손 다건 정본(지점장 확정 2026.07.18): 실손은 최대 3건 —
    #   '상해의료비 + 실손 + 상해실손' 조합. 상해의료비 행은 '실손' 구분 밖에 있어
    #   기존 _sil_rows(=실손 그룹 내부만)로는 잡히지 않아 1건으로 뭉개졌다.
    #   → 시트 전체에서 '상해의료비'/'상해실손' 행을 찾아 실손 판정행에 추가한다.
    for _r in range(1, ws.max_row + 1):
        _b = str(ws.cell(_r, 2).value or '')
        if ('상해의료비' in _b or '상해실손' in _b) and _r not in _sil_rows:
            _sil_rows.append(_r)
    # 헤더(계약 메타)
    headers=[]
    # ★★★★★v419 (지점장 지적 2026.08.13 「2번페이지도 새로가입하는건에대해 다 블랙이다」)
    #   구 루프는 `range(2,last)`라 <b>합산 열(보유 합계·제안 합계)까지 계약으로 셌다</b>.
    #   실측: 계약 21건(실제 19건) · 보험료 막대에 「보유 2,852,227원」「제안 119,572원」 막대가
    #   계약인 척 서 있었다 · 표지도 21건. v417은 <b>색만</b> 고치고 계약 판정은 그대로였다 — 내 누락.
    #   ★<b>「계약이냐」는 판정은 한 곳에서만 한다</b>: 헤더 1행이 합산 라벨인 열은 계약이 아니다.
    _SUMHDR=('보유 합계','제안 합계','합계','보유합계','제안합계')
    _propcols=set()
    for _c in range(2,last+1):
        _h1=str(ws.cell(1,_c).value or '').strip()
        if _h1 in _SUMHDR and _h1.replace(' ','')=='제안합계':
            # 제안 합계 열 <b>왼쪽</b>이 제안 계약 열이다(엑셀 합산 2열 구조 v388c)
            for _p in range(_c-1,1,-1):
                if str(ws.cell(1,_p).value or '').strip() in _SUMHDR: continue
                _propcols.add(_p); break
    for c in range(2,last):  # 마지막=합계 제외
        nm=ws.cell(1,c).value; pr=_man(ws.cell(2,c).value)
        if str(nm or '').strip() in _SUMHDR: continue     # ★v419 합산 열은 계약이 아니다
        if nm is None and pr==0: continue
        _raw=str(nm or '')                                       # ★원본(회사\n상품\n[갱신])
        _rl=[x.strip() for x in _raw.split('\n') if x.strip()]
        _co_=(_rl[0] if _rl else '')
        _pr_=(_rl[1] if len(_rl)>1 else '')
        _pr_=re.sub(r'\s*\[[^\]]*\]','',_pr_).strip()
        nm=str(nm or '').replace('\n',' ').strip()
        renew = '[갱신]' in nm   # 헤더 형식 [갱신]/[비갱신(종신)] — 대괄호 정확매칭
        nm=re.sub(r'\s*\[[^\]]*\]','',nm)                       # [갱신]·[비갱신(종신)] 제거
        nm=re.sub(r'\((무|표준형|종신|표준형-종신|갱신|비갱신)\)','',nm)  # 괄호 수식 제거
        nm=re.sub(r'\s+',' ',nm).strip()
        _join=str(ws.cell(3,c).value or '').strip()             # 3행=가입년일
        _hassil=any(_man(ws.cell(r,c).value)>0 for r in _sil_rows)  # 실손 담보 보유 계약?
        # ★v89 실손 체크 보강(지점장 2026.07.19): 담보값이 비어 있어도 회사·상품명에
        #   '실손'·'의료비'가 적혀 있으면 실손 계약으로 체크한다(1+1=2 — 2건이면 2건).
        if not _hassil:
            _tag=(str(_co_ or '')+' '+str(_pr_ or ''))
            if ('실손' in _tag) or ('의료비' in _tag):
                _hassil=True
        # ★v79 단체보험 제외(지점장 확정 2026.07.18): ①가입~만기 1년 ②상품명에 '단체' — 2조건 동시 충족만 제외
        try:
            _pp = str(_pr_ or '').replace(' ', '')
            if '단체' in _pp:
                _cy = re.match(r'(\d{4})', str(_join or ''))
                _ey = re.match(r'(\d{4})', str(ws.cell(4, c).value or ''))
                if _cy and _ey and (int(_ey.group(1)) - int(_cy.group(1))) <= 1:
                    continue      # 단체보험 → 계약열 자체를 제외
        except Exception:
            pass
        # ★v89: 헤더에 '(3세대 실손)'처럼 세대가 적혀 있으면 힌트로 보관(가입일이 비었을 때 사용)
        _gh=re.search(r'\((\d)\s*세대', _raw)
        _np3=any(str(ws.cell(r,c).value or '').strip() not in ('','0') for r in getattr(load_excel,'_np3_rows',[]))
        _drug=any(str(ws.cell(r,c).value or '').strip() not in ('','0') for r in getattr(load_excel,'_drug_rows',[]))
        # ★★★★★v463 제71조 (지점장 지적 2026.08.17
        #   「보장진단서 12페이지에 종신사망 계약리스트가 기재가 안 된다」)
        #   실측: 12쪽은 `rep['contracts']`를 읽는데 그 키가 <b>아예 없었다</b>(None) →
        #   연금·종신·저축이 전부 「미보유」. 판정 함수는 멀쩡했고 <b>데이터가 도달을 못 했다</b>.
        #   → 헤더 원본을 그대로 실어 보낸다. 종신 판정은 헤더의 `[비갱신(종신)]`이 원천이다.
        #     실측: 교보생명 / 교보3밸런스보장보험 (무배당) / [비갱신(종신)]
        headers.append({'nm':nm or '계약','amt':int(pr),'renew':renew,'join':_join,'sil':_hassil,
                        'raw_hdr':_raw,                               # ★v463 헤더 원본(3줄)
                        'company':_co_, 'product':_pr_,               # ★v463 12쪽이 쓰는 이름
                        'renewal':(re.search(r'\[[^\]]*\]', _raw).group(0) if re.search(r'\[[^\]]*\]', _raw) else ''),
                        'contract_date':_join,
                        'premium':int(pr),
                        'lump_sum':(int(pr) if '일시납' in str(ws.cell(2,c).value or '') else 0),
                        'pay_term':str(ws.cell(5,c).value or '').strip(),
                        'prop':(c in _propcols),                      # ★v419 가입제안서 계약 = 레드
                        'co':_co_,'prod':_pr_,'np3':_np3,'drug':_drug,
                        'genhint':(int(_gh.group(1)) if _gh else None)})
    total_prem=int(_man(ws.cell(2,last).value))
    return grp_rows, headers, total_prem

def cat_total(grp_rows, cat):
    """리포트 카테고리 보유합(만원) + 대표담보 top3"""
    rows=[]
    for g in CATEGORY_GROUPS[cat]:
        rows+=grp_rows.get(g,[])
    total=sum(v for _,v in rows)
    top=sorted([(b,v) for b,v in rows if v>0], key=lambda x:-x[1])[:3]
    # ★★v95 (지점장 지시 2026.07.19): <b>상해의료비는 보장진단서에 반드시 표기</b>한다.
    #   금액이 작아 top3에서 잘려 사라지던 것을 고정 노출로 바꾼다(실손과 별개 담보).
    _pin=[(b,v) for b,v in rows if v>0 and '상해의료비' in str(b)]
    if _pin and not any('상해의료비' in str(b) for b,_ in top):
        top=(top[:2] if len(top)>2 else top)+_pin[:1]
    return total, top, rows

def pct_for(cat, grp_rows, band):
    total,top,rows=cat_total(grp_rows,cat)
    if cat in BENCHMARK[band]:
        rec=BENCHMARK[band][cat]
        return (round(total/rec*100) if rec else 0), total, top
    # presence 계열
    spec=PRESENCE.get(cat)
    if not spec: return 0,total,top
    have=sum(1 for b,v in rows if v>0 and any(k in b for k in spec['keys']))
    return round(have/spec['need']*100), total, top

def map_excel_to_report(xlsx_path, settings=None, age_band='40s', age_known=False):
    """완성 엑셀 → rep dict (report_weasy.build_report_pdf 입력)"""
    settings=settings or {}
    grp_rows, headers, total_prem = load_excel(xlsx_path)
    # ★v134 흥국화재 10억통장 가입 판정(지점장 확정 2026.07.21):
    #   엑셀 헤더 1행(회사\n상품\n[갱신])에 <b>'흥국' AND '리셋월렛'</b>이 둘 다 있으면 가입.
    # ★★★v139 갱신 색 원천(지점장 2026.07.21): 엑셀 값 글자색이 파랑(0070C0)이면 그 담보는 '갱신'.
    #   엑셀이 이미 정본이므로 색을 그대로 읽어 설명서·PPT에 전달한다(4대 산출물 연동).
    _gen_map={}; _red_map={}   # ★v370 _red_map = 가입제안서(레드 C00000) 담보
    _own_amt={}; _prop_amt={}  # ★v417 담보별 보유합계 / 제안합계(2줄 분리 표기용)
    try:
        import openpyxl as _ox2
        _w2=_ox2.load_workbook(xlsx_path); _s2=_w2.active
        # ★★★★★v417 (지점장 지적 2026.08.13 「6페이지는 다 레드로 나온다」) — 근본 원인:
        #   구 코드는 `range(3, max_column)`으로 <b>합산 열까지 훑었다</b>. 엑셀 합산 라인 2열(v388c)의
        #   <b>제안 합계 열은 글자색이 레드 C00000이고 값이 `=SUM(...)` 수식 문자열</b>이라
        #   <b>모든 담보 행에서 항상 참</b>이 된다 → `_red_map`에 <b>담보 99개 전부</b>가 들어가
        #   설명서·진단서가 <b>통째로 레드</b>로 찍혔다(실측 99/99).
        #   → <b>색은 실제 계약 열에서만 읽는다.</b> 합산 3열(보유 합계·제안 합계·합계)은 제외.
        _SUMHDR=('보유 합계','제안 합계','합계','보유합계','제안합계')
        _datac=[_c for _c in range(3,_s2.max_column+1)
                if str(_s2.cell(1,_c).value or '').strip() not in _SUMHDR]
        _sumc={}
        for _c in range(3,_s2.max_column+1):
            _h=str(_s2.cell(1,_c).value or '').strip()
            if _h in _SUMHDR: _sumc[_h.replace(' ','')]=_c
        for _r2 in range(6,_s2.max_row+1):
            _nm2=_s2.cell(_r2,2).value
            if not _nm2: continue
            _nm2=str(_nm2).strip()
            for _c2 in _datac:
                _f2=_s2.cell(_r2,_c2).font
                _rgb=(_f2.color.rgb if (_f2 and _f2.color and _f2.color.rgb) else '')
                _up2=str(_rgb).upper() if _rgb else ''
                _hasv=_s2.cell(_r2,_c2).value not in (None,'',0)
                if _up2.endswith('C00000') and _hasv: _red_map[_nm2]=True   # ★v370 가입제안서
                if _up2.endswith('0070C0') and _hasv:
                    _gen_map[_nm2]=True
        _w2.close()
        # ★v417 보유/제안 금액 분리(캐시값) — 진단서 2줄 표기의 유일 원천은 엑셀이다(결과값 동결 #9).
        if _sumc.get('보유합계') and _sumc.get('제안합계'):
            _w3=_ox2.load_workbook(xlsx_path, data_only=True); _s3=_w3.active
            for _r3 in range(6,_s3.max_row+1):
                _nm3=_s3.cell(_r3,2).value
                if not _nm3: continue
                _nm3=str(_nm3).strip()
                _ov=_s3.cell(_r3,_sumc['보유합계']).value
                _pv=_s3.cell(_r3,_sumc['제안합계']).value
                if isinstance(_ov,(int,float)) and _ov: _own_amt[_nm3]=_ov
                if isinstance(_pv,(int,float)) and _pv: _prop_amt[_nm3]=_pv
            _w3.close()
        print(f'[v417 색원천] 계약열 {len(_datac)}개 · 합산열 제외 {len(_sumc)}개 · red_map {len(_red_map)} · gen_map {len(_gen_map)} · 제안금액 {len(_prop_amt)}건')
    except Exception as _e417:
        print(f'[v417 색원천] 실패 {_e417}')
    # ★★★v182 (지점장 2026.07.22): 세부가입현황 미파싱 등 <b>수기 확인이 필요한 건</b>을
    #   확인사항 시트에서 모아 보장진단서에 <b>빨간 경고 배너</b>로 띄운다(방치 출고 차단).
    _warn=[]
    try:
        import openpyxl as _ox3
        _w3=_ox3.load_workbook(xlsx_path)
        if '확인사항' in _w3.sheetnames:
            _s3=_w3['확인사항']
            for _r3 in range(1,_s3.max_row+1):
                _t3=' '.join(str(_s3.cell(_r3,_c3).value or '') for _c3 in range(1,6))
                if '세부가입현황' in _t3 or '축 대조' in _t3 or '상세내역' in _t3:
                    _co3=str(_s3.cell(_r3,1).value or '').strip()
                    _nm3=str(_s3.cell(_r3,2).value or '').strip()
                    # ★v183 팩트만 남긴다: 설명문 제거 → 담보명 · 배치결과
                    _res3 = '뇌출혈로 배치' if '뇌출혈로 배치' in _nm3 else '축 미확정'
                    _dm3 = re.sub(r'^\[확인\]\s*', '', _nm3)
                    _dm3 = re.sub(r'세부가입현황[^·]*?(요망|필수)\s*', '', _dm3).strip()
                    _dm3 = re.sub(r'^[·\s]+', '', _dm3)[:28] or _nm3[:28]
                    _warn.append((_co3, _dm3, _res3))
        _w3.close()
    except Exception: pass
    # ★★v185 (지점장 2026.07.22): AIA·AIG·라이나(우체국) 계약이 <b>하나라도 있으면</b>
    #   파싱 성공 여부와 무관하게 <b>"뇌 범위 부분 꼭 체크"</b> 경고를 띄운다(1000% 대조 원칙).
    _warn_co=[]
    try:
        import openpyxl as _ox4
        _w4=_ox4.load_workbook(xlsx_path); _s4=_w4.active
        for _c4 in range(3,_s4.max_column+1):
            _h4=str(_s4.cell(1,_c4).value or '').replace(' ','')
            for _f4 in ('AIA','AIG','라이나','우체국'):
                if _f4 in _h4 and _f4 not in _warn_co: _warn_co.append(_f4)
        _w4.close()
    except Exception: pass
    _r10=bool(settings.get('reset10')) if settings.get('reset10') is not None else False
    try:
        if settings.get('reset10') is not None: raise StopIteration
        import openpyxl as _ox
        _w=_ox.load_workbook(xlsx_path); _s=_w.active
        for _c in range(2,_s.max_column+1):
            _t=str(_s.cell(1,_c).value or '').replace(' ','')
            if ('흥국' in _t) and ('리셋월렛' in _t or '리셋월랫' in _t): _r10=True; break
        _w.close()
    except Exception: pass
    _badj=_bundle_adjust(xlsx_path)   # ★v30e 묶음 전개 중복 차감(심장·뇌 보유합)
    client=settings.get('client','고객')

    # ★★★★★v418 (지점장 지시 2026.08.13 「담보이름으로」) — 충족률 「보유」를 <b>담보명 기준</b>으로.
    #   구 방식은 A열 <b>구분 그룹의 세로합</b>이었다. 이영태 뇌혈관 8,000만 =
    #   뇌혈관진단비 2,500 + 뇌졸증진단비 1,500 + <b>중대한 뇌졸증(CI) 4,000</b>.
    #   ★<b>세로로 더한 숫자는 고객이 한 번에 받는 돈이 아니다.</b> 지점장 원문 「뇌혈관 다 해도 2천인데」.
    #   CI는 중대성 요건을 통과해야 나오는 별개 담보라 표준 컬럼에 더하지 않는다(지침 원칙).
    #   ★담보가 2개면 <b>낮은 쪽</b>을 본다 — 충족률은 「어디가 비었나」를 보는 지표다.
    def _amt_of(nm):
        for rows in grp_rows.values():
            for b,v in rows:
                if str(b).strip()==nm: return v or 0
        return 0
    def _named(cat):
        spec=NAMED_BENCH.get(cat)
        if not spec: return None
        out=[]
        for nm,rec in spec:
            a=_amt_of(nm)
            out.append({'nm':nm,'amt':a,'rec':rec,'pct':(round(a/rec*100) if rec else 0)})
        # ★★★★★v421 (지점장 지적 2026.08.14 박미정 「5페이지 심장 0%인거 오류」)
        #   v418의 「담보 2개면 낮은 쪽」이 <b>미가입(0)인 담보</b>를 골라 항상 0%가 됐다.
        #   실측: 허혈성 2,000만이 엑셀에 있는데 급성심근경색(0)이 기준이 되어 <b>심장 0%</b>.
        #   → 제1조(엑셀이 담보의 전부다)·제2조(결과값 동결) 위반. 엑셀에 있는데 화면은 0이다.
        #   ★<b>가입된 담보 중에서</b> 낮은 쪽을 본다. 전부 미가입일 때만 0%.
        _live=[x for x in out if x['amt']]
        return min(_live or out, key=lambda x:x['pct'])
    coverage=[]; donut_map={}; detail_map={}; named_map={}
    for cat in CATEGORY_GROUPS:
        p,total,top=pct_for(cat,grp_rows,age_band)
        _nb=_named(cat)
        if _nb:
            p=_nb['pct']; named_map[cat]=_nb
            donut_map[cat]=p
            detail_map[cat]={'have':(_fmt(_nb['amt']) or '0'),'rec':_fmt(_nb['rec']),
                             'unit':'만','base':_nb['nm']}
            status='full' if p>=70 else ('part' if p>=40 else 'gap')
            _disp=getattr(load_excel,'_disp',{})
            items=[{'t':b,'v':(_disp.get(b) or _fmt(v)),
                    **({'blue':True} if (cat in ('실손·일배책',) or _gen_map.get(str(b).strip())) else {}),
                    **({'red':True} if _red_map.get(str(b).strip()) else {})} for b,v in top]
            if not items or all(not it['v'] for it in items):
                items=[{'t':f'{cat} 없음','none':True}]
            coverage.append({'name':cat if cat!='심장' else '심장 (＋빈맥)','status':status,'items':items})
            continue
        if cat in ('심장','뇌혈관') and _badj.get(cat):
            total=max(0,total-_badj[cat])
            rec=BENCHMARK[age_band].get(cat)
            if rec: p=round(total/rec*100)
        donut_map[cat]=p
        if cat in BENCHMARK[age_band]:
            detail_map[cat]={'have':_fmt(total) or '0','rec':_fmt(BENCHMARK[age_band][cat]),'unit':'만'}
        else:
            spec=PRESENCE.get(cat,{'keys':[],'need':1})
            have=sum(1 for b,v in cat_total(grp_rows,cat)[2] if v>0 and any(k in b for k in spec['keys']))
            detail_map[cat]={'have':f'{have}개','rec':f"{spec['need']}개",'unit':'개'}
        status='full' if p>=70 else ('part' if p>=40 else 'gap')
        # ★★★v210 (지점장 확정 2026.07.25, 영구): <b>'입원·일당' 카테고리 강제 파랑 폐기</b>.
        #   간병인 · 간호통합병동을 포함한 일당 계열은 보험료 · 가입년일 · 만기일자 · 총납입기간(=계약 갱신 판정)
        #   또는 담보명 [갱신] 표기에 따라 <b>엑셀 글자색(gen_map)을 그대로 따라간다</b>. 실손 · 일배책만 항상 파랑.
        blue = cat in ('실손·일배책',)
        _disp=getattr(load_excel,'_disp',{})   # ★v30h 슬래시 원문 우선
        # ★★★v147 (지점장 지적 2026.07.21): 파랑이 <b>카테고리 단위</b>로만 걸려 있어
        #   갱신 담보인 '허혈성 진단비'(엑셀 글자색 0070C0)가 보장현황에서 검정으로 나왔다.
        #   → 담보별 gen_map(엑셀 글자색=원천)을 함께 본다. 4대 산출물 색 연동.
        items=[{'t':b,'v':(_disp.get(b) or _fmt(v)),
                **({'blue':True} if (blue or _gen_map.get(str(b).strip())) else {}),
         **({'red':True} if _red_map.get(str(b).strip()) else {})} for b,v in top]
        if not items or all(not it['v'] for it in items):
            items=[{'t':f'{cat} 없음','none':True}]
        coverage.append({'name':cat if cat!='심장' else '심장 (＋빈맥)','status':status,'items':items})

    # 강점/공백 = pct 임계
    ranked=sorted(donut_map.items(), key=lambda x:-x[1])
    strength=[{'h':c,'d':f'충족률 {p}% — 핵심담보 보유'} for c,p in ranked if p>=70][:4]
    weak=[{'h':c,'d':f'충족률 {p}% — 보강 필요'} for c,p in ranked if p<40][:4]
    gap_count=sum(1 for _,p in ranked if p<40)

    # ★v419 제안 계약은 목록·막대에서 레드(C00000)로 구분한다 — 보유와 섞이면 「새로 가입하는 건」이 안 보인다
    renew_list=[{'nm':h['nm'][:18],'v':f"{h['amt']:,}원",**({'prop':True} if h.get('prop') else {})}
                for h in headers if h['renew']]
    nonren_list=[{'nm':h['nm'][:18],'v':f"{h['amt']:,}원",**({'prop':True} if h.get('prop') else {})}
                 for h in headers if not h['renew']]
    _co=lambda nm:(nm.split(' ')[0] if ' ' in nm else nm[:6])
    bars=sorted([{'nm':_co(h['nm']),'amt':h['amt'],'renew':h['renew'],'prop':bool(h.get('prop'))} for h in headers],
                key=lambda x:-x['amt'])
    donuts=[{'name':('심장' if c=='심장' else c.split('·')[0] if c in('실손·일배책','입원·일당','응급실·독감','골절·화상','사망·후유') else c),
             'pct':min(100,donut_map[c])} for c in DONUT_ORDER]
    # 도넛 라벨 보정
    label={'실손·일배책':'실손·배상','입원·일당':'입원·일당','응급실·독감':'응급실·독감','골절·화상':'골절·화상','사망·후유':'사망·후유'}
    # ★버킷 방식(지점장 2026.07.12): 도넛은 100%까지만, 초과분은 '충분+' 배지
    _raw={}
    for c in DONUT_ORDER:
        t,_tp,_rw=cat_total(grp_rows,c)
        if c in ('심장','뇌혈관') and _badj.get(c): t=max(0,t-_badj[c])
        if c in BENCHMARK[age_band] and BENCHMARK[age_band][c]:
            _raw[c]=round(t/BENCHMARK[age_band][c]*100)
        else:
            _sp=PRESENCE.get(c,{'keys':[],'need':1})
            _hv=sum(1 for b,v in _rw if v>0 and any(k in b for k in _sp['keys']))
            _raw[c]=round(_hv/_sp['need']*100) if _sp['need'] else 0
    donuts=[{'name':label.get(c,c),'pct':min(100,donut_map[c]),'raw':_raw.get(c,0),
             'over':_raw.get(c,0)>100} for c in DONUT_ORDER]

    # ── 치료비 정리 5항목 (라벨 → 엑셀 정확 담보명) ──
    CHIRYO=[('암주요치료비','암주요치료비'),('비급여주요치료비','하이클래스(암)'),
            ('순환계주요치료비','2대 주요치료비'),('산정특례(뇌혈관)','산정특례뇌혈관'),
            ('산정특례(심장)','산정특례심장')]
    allrows={}
    for rows in grp_rows.values():
        for b,v in rows:
            if v>0: allrows[b]=max(allrows.get(b,0),v)
    # ★v39 원본담보명 로드 (_dambo_raw 숨김시트) → 워크시트 흰칸에 '담보명 금액' 표기
    _raw_map={}
    try:
        import openpyxl as _ox
        _wbr=_ox.load_workbook(xlsx_path, data_only=True)
        if '_dambo_raw' in _wbr.sheetnames:
            _rs=_wbr['_dambo_raw']
            for _r in range(2,_rs.max_row+1):
                _s=_rs.cell(_r,1).value; _rw=_rs.cell(_r,2).value
                if _s and _rw: _raw_map[str(_s).strip()]=str(_rw).strip()
    except Exception:
        pass
    def _clean_dnm(nm):   # 원본담보명 정리: 괄호주석·병원접두·회차 제거해 짧게
        import re as _re
        s=_re.sub(r'[（(].*?[）)]','',nm)                       # 괄호내용 제거
        s=_re.sub(r'상급종합병원[ᅵI|\s]*(II|Ⅱ|III|Ⅲ)?\s*','',s) # 병원접두 제거
        s=_re.sub(r'\s+',' ',s).strip(' ·-|Ⅲ III II')
        return s[:22]
    chiryo=[{'name':lab,'value':(_fmt(allrows.get(key,0)) or '미가입'),
             'raw':_clean_dnm(_raw_map[key]) if _raw_map.get(key) else ''} for lab,key in CHIRYO]

    # ── CI/생명보험 선지급 분석 (사망값 의존X: 선지급률=본체/(본체+잔여)) ──
    def _gv(nm):
        for rows in grp_rows.values():
            for b,v in rows:
                if str(b).strip()==nm: return v
        return 0
    # ★P5 상단 핵심 진단비 3종(지점장 지시): 뇌출혈·뇌졸증·급성심근경색
    def _fv(v): return _fmt(v) if v and v>0 else '미가입'
    # ★★★★★v270 영구지침(지점장 지적 2026.07.30 "진단서에 없는 뇌출혈 담보가 찍혀있다"):
    #   구 코드는 뇌출혈 칸 폴백이 <b>`or _gv('중대한 뇌졸증')`</b>까지 내려가서,
    #   뇌출혈 담보가 하나도 없는 고객에게 <b>축이 다른 뇌졸증 값을 뇌출혈 칸에 찍었다</b>.
    #   실측(이명순): 엑셀 뇌출혈진단비 0 · 중대한 뇌출혈 0 인데 진단서 8p에 <b>4,000만</b>이 인쇄됐다(등식1 위반).
    #   → <b>뇌출혈 칸은 뇌출혈 축(뇌출혈진단비 · 중대한 뇌출혈)만 본다.</b> 없으면 '미가입'.
    #   ★뇌 축 규칙(§CI 5b)은 '축을 고른다'는 뜻이지 <b>다른 축 값을 복사하라는 뜻이 아니다</b>.
    _p5=[('뇌출혈진단비',_gv('뇌출혈진단비') or _gv('중대한 뇌출혈')),
         ('뇌졸증진단비',_gv('뇌졸증진단비') or _gv('뇌혈관진단비')),
         ('급성심근경색',_gv('급성심근경색') or _gv('중대한 급성심근') or _gv('허혈성 진단비')),
         ('뇌혈관진단비',_gv('뇌혈관진단비')),
         ('허혈성 진단비',_gv('허혈성 진단비') or _gv('허혈성심장질환'))]
    p5_own=[{'t':n,'v':_fv(v)} for n,v in _p5]
    # ★v43 산정특례 금액 맵(6p 보장나무 유동 표시) — 보유 시에만 값 존재
    spec_amounts={}
    _sb=_gv('산정특례뇌혈관'); _sh=_gv('산정특례심장')
    if _sb>0: spec_amounts['brain']=_fv(_sb)
    if _sh>0: spec_amounts['heart']=_fv(_sh)
    _ci_pairs=[(n,_gv(n)) for n in ('중대한 암','중대한 뇌졸증','중대한 뇌출혈','중대한 급성심근')]
    _ci_pairs=[(n,v) for n,v in _ci_pairs if v>0]
    _ci_apply=_gv('중대한CI적용')
    _ci_bonche=max((v for _,v in _ci_pairs), default=0)
    # ★v33 선지급률: 끝열(_ci_apply)은 비CI 일반사망 오염 → CI 계약 열에서 직접 산출
    # ★★★v236: 계약별 CI 리스트(최대 3) — 3p를 계약 수만큼 분리 표기하기 위한 원천
    _ci_list=_ci_meta_list(xlsx_path)
    _cm=_ci_meta(xlsx_path)
    if _cm:
        _ci_bonche=_cm['bonche']; _ci_samang=_cm['samang']; _ci_rate=_cm['pct']; _ci_apply=_cm['resid']
    else:
        _ci_samang=_ci_bonche+_ci_apply
        _ci_rate=round(_ci_bonche/_ci_samang*100) if _ci_samang else 0
        _ci_rate=(80 if abs(_ci_rate-80)<=abs(_ci_rate-50) else 50) if _ci_rate else 0
    ci={'present':bool(_ci_pairs or _ci_apply>0),'samang':_fmt(_ci_samang),
        'rate':_ci_rate,'residual':_fmt(_ci_apply),
        'list':_ci_list,'n_ci':len(_ci_list),
        'items':[{'t':{'중대한 암':'ci암진단비','중대한 뇌졸증':'ci뇌졸증','중대한 급성심근':'ci급성심근경색','중대한 뇌출혈':'ci뇌출혈'}.get(n,n),'v':_fmt(v)} for n,v in _ci_pairs]}
    # ★CI 3상태 판정(2026.07.07 지점장): 상품명 CI/GI/리빙케어 + 중대한OO담보 값
    def _ciprod1(nm):
        # ★v149 위 _isci와 동일 기준(퍼펙트플러스·퍼펙트통합 포함). 정본 1개로 통일.
        t=re.sub(r'[\s\u3000]','',str(nm or ''))
        # ★★★v235 (한정환 실측): 구 `'CI보험'` 연속매칭이 `CI종신보험`을 놓쳤다(CI와 '보험' 사이에
        #   '종신' 등 수식어가 끼면 탈락). main.py `_isci_prod`와 <b>정본 1개로 통일</b>.
        #   영문 경계 필수 — `ACCIDENT` 안의 우연한 'CI' 배제.
        if ('퍼펙트' in t) or ('퍼텍트' in t) or ('리빙케어' in t): return True
        return bool(re.search(r'(?<![A-Za-z])(CI|GI)(?![A-Za-z])', t)) and ('보험' in t or '종신' in t)
    _ci_prod=any(_ciprod1(h.get('nm','')) for h in headers)
    # ★2026.07.12 지점장 확정: 상품명에 CI/GI/리빙케어가 없으면 '중대한OO' 담보가 있어도 진짜 CI가 아니다(가짜).
    #   → 상품명이 1순위. 상품명에 표기 없으면 무조건 none(Plan B).
    if not _ci_prod:
        ci['status']='none'        # 상품명에 CI/GI/리빙케어 없음 = 확실 비CI (중대한OO는 가짜)
        ci['present']=False
        ci['items']=[]
    elif _ci_pairs or _ci_apply>0:
        ci['status']='ci'          # 상품명 CI + 담보값 있음 = 확실 CI
    else:
        ci['status']='check'       # 상품명 CI인데 담보값 없음/애매 = 회색지대 [확인]

    # ── Plan B: 비CI 진단비 정액 지급 구조 (CI 미보유 시 P3 상단 CI블록 대체) ──
    def _sumnm(*names):
        s=0
        for rows in grp_rows.values():
            for b,v in rows:
                if str(b).strip() in names and v>0: s+=v
        return s
    _amt_cancer=max(_gv('일반암'),_gv('고액암'))
    _amt_brain=_sumnm('뇌혈관진단비','뇌졸증진단비')
    _amt_heart=_sumnm('급성심근경색','허혈성 진단비')
    noci_items=[]
    if _amt_cancer>0: noci_items.append({'t':'암 진단비','v':_fmt(_amt_cancer)})
    if _amt_brain>0:  noci_items.append({'t':'뇌혈관·뇌졸증','v':_fmt(_amt_brain)})
    if _amt_heart>0:  noci_items.append({'t':'급성심근·허혈성','v':_fmt(_amt_heart)})
    noci={'present':(not ci['present']) and bool(noci_items),'items':noci_items}

    rep={
        'client':client,
        'branch':settings.get('branch',''),'manager':settings.get('manager',''),
        'title':settings.get('title',''),'phone':settings.get('phone',''),
        'contracts':headers,   # ★v463 제71조 — 12쪽 재무 표가 읽는 계약 목록(없어서 「미보유」였다)
        'n_contract':len(headers),'premium':total_prem,'reset10':_r10,'reset10_amt':settings.get('reset10_amt',0),'gen_map':_gen_map,'red_map':_red_map,'own_amt':_own_amt,'prop_amt':_prop_amt,'warn_list':_warn,'warn_co':_warn_co,
        'renew':len(renew_list),'nonrenew':len(nonren_list),'gap_count':gap_count,
        'coverage':coverage,'strength':strength,'weak':weak,
        'renew_list':renew_list,'nonrenew_list':nonren_list,
        'premium_bars':bars,'donuts':donuts,
        'donut_detail':[{'name':label.get(c,c),'have':detail_map[c]['have'],
                         'base':detail_map[c].get('base',''),
                         # ★v418 표는 <b>실제치</b>(주석 「상한 없음·실제치, 2026.07.12 지점장 확정」 복원).
                         #   도넛 링은 물리적으로 100%가 끝이라 링만 min(100)을 유지한다.
                         'rec':detail_map[c]['rec'],'pct':donut_map[c],
                         'raw':_raw.get(c,0),'over':_raw.get(c,0)>100} for c in DONUT_ORDER],
        'band_label':{'20s':'20대','30s':'30대','40s':'40대','50s':'50대','60s':'60대'}.get(age_band,age_band),
        'chiryo':chiryo,
        'ci':ci,
        'noci':noci,
        'p5_own':p5_own,
        'spec_amounts':spec_amounts,
        'age_band':age_band,'age_known':age_known,
    }
    # ── 리모델링 제안: 1-5종 권유 · 운전자 재가입 권유 (지침 §7·§8.6) ──
    advice=[]
    try:
        _wb2=openpyxl.load_workbook(xlsx_path, data_only=True); _ws2=_wb2.active
        _rowmap={}
        for _r in range(6,_ws2.max_row+1):
            _b=_ws2.cell(_r,2).value
            if _b: _rowmap[str(_b).strip()]=_ws2.cell(_r,_ws2.max_column).value
        def _has(nm):
            v=_rowmap.get(nm); return (v not in (None,'',0)) and (str(v).strip() not in ('','0'))
        def _num(nm):
            v=_rowmap.get(nm)
            try: return int(float(str(v).replace(',','').split('/')[0]))
            except: return 0
        # 수술: 1-8종/N대만 있고 1-5종 없으면 → 1-5종 권유
        if (_has('상해 종수술비(1-8종)') or _has('질병 종수술비(1-8종)') or _has('120대수술비')) \
           and not (_has('상해 종수술비(1-5종)') or _has('질병 종수술비(1-5종)')):
            advice.append({'t':'수술비 리모델링 권유',
                'd':'현재 수술비가 1-5종이 아닙니다(1-7·8·9종 또는 N대). 이 유형은 비급여 항목 미보장·청구 절차 복잡으로 실질 보장 범위가 좁습니다. 관혈/비관혈을 구분하고 비급여까지 보장하는 1-5종 수술비 가입(전환)을 권유드립니다.'})
        # 운전자: 최신 기준 미달이면 → 재가입 권유
        _std={'합의금':20000,'6주미만':1000,'대인':3000,'대물':500,'변호사':5000}
        _lab={'합의금':'합의금(기준 2억)','6주미만':'6주미만 합의금(기준 1천만)','대인':'벌금 대인(기준 3천만)','대물':'벌금 대물(기준 500만)','변호사':'변호사선임(기준 5천만)'}
        if any(_num(k)>0 for k in _std):
            _short=[f"{_lab[k]}: {('보유 '+format(_num(k),',')+'만') if _num(k)>0 else '미보유'}" for k in _std if _num(k)<_std[k]]
            if _short:
                advice.append({'t':'운전자 리모델링 권유',
                    'd':'2022년 보행자보호의무·처벌 강화로 옛 운전자보험은 담보가 부족합니다. 미달 항목 — '+' / '.join(_short)+'. 최신 기준으로 재가입(리모델링)을 권유드립니다.'})
    except Exception:
        pass
    # ── ★P5 질병코드 커버표: 고객 실제 보유 감지 (하드코딩 제거) ──
    _allnm=set()
    for _rows in grp_rows.values():
        for _b,_v in _rows:
            if _v and _v>0: _allnm.add(str(_b).strip())
    def _any(*subs):
        return any(any(s in nm for s in subs) for nm in _allnm)
    scope_brain=[]   # 보유 행 key
    # ★★★★★v399 (지점장 지적 2026.08.12): <b>「보유」는 그 이름의 담보를 실제로 가졌을 때만</b>.
    #   지점장 원문: 「<b>여긴 뇌졸증이 없는데 왜 계속 뇌졸증이 나오냐 진단서에</b>」
    #   ★실측: 사공호는 KB `뇌혈관질환진단비Ⅲ` 하나뿐인데 6p 질병코드표의
    #     <b>뇌출혈·뇌졸증·기타뇌혈관 세 행이 전부 노란행+「보유」</b>로 찍혔다.
    #     구 조건이 `뇌혈관진단`을 hem·infarct에도 넣었기 때문이다.
    #   → <b>뇌출혈·뇌졸증 행은 그 이름의 담보가 있을 때만 보유</b>.
    #     `other`(기타 뇌혈관질환)는 뇌혈관진단비 <b>그 자체</b>이므로 종전대로 둔다.
    #   ★심장 쪽은 지시 범위 밖 — 손대지 않았다(확인 대기).
    if _any('뇌출혈','중대한 뇌졸증'): scope_brain.append('hem')
    if _any('뇌경색','뇌졸증진단'): scope_brain.append('infarct')
    if _any('뇌혈관진단'): scope_brain.append('other')
    if _any('외상성뇌출혈','외상성 뇌출혈'): scope_brain.append('trauma')
    if _any('산정특례뇌','산정특례(뇌'): scope_brain.append('brain_snjt')
    scope_heart=[]
    if _any('급성심근','중대한 급성심근'): scope_heart.append('ami')
    # ★★★★★v402 (지점장 확정 2026.08.12, 영구): <b>허혈성은 단독이다</b>.
    #   지점장 원문: 「<b>허헐성은 허혈성단독이라고</b>」 — 제「단독 5종」 조문과 동일.
    #   ★구 코드는 허혈성진단비 하나로 <b>협심증(angina)</b>까지 보유로 찍었다.
    #     제25조(「보유는 그 이름의 담보가 있을 때만」)를 심장에도 그대로 적용한다.
    #   → <b>협심증 행은 협심증 담보(또는 묶음 분해분)가 있을 때만</b> 보유.
    #     `chronic`(기타·만성 허혈 I24·25)은 <b>허혈성진단비 그 자체</b>이므로 종전대로 둔다
    #     (뇌의 `other`(기타 뇌혈관질환)와 같은 자리).
    if _any('허혈성 진단비','허혈심장','허혈성진단'): scope_heart.append('chronic')
    if _any('협심증'): scope_heart.append('angina')      # ★v50 심장 묶음 분해분(단독 보유)
    if _any('부정맥'): scope_heart.append('arrhy')
    if _any('심부전'): scope_heart.append('hf')
    if _any('심장판막','판막'): scope_heart.append('valve')
    if _any('심장염증','심근염','심내막','심장막','염증'): scope_heart.append('inflam')
    if _any('심근병증'): scope_heart.append('cardiomyo')
    if _any('산정특례심장','산정특례(심'): scope_heart.append('heart_snjt')
    rep['scope_brain']=scope_brain
    scope_heart=list(dict.fromkeys(scope_heart))   # ★v50 중복 제거
    rep['scope_heart']=scope_heart
    rep['advice']=advice
    # ★ 전체 담보→표시값 맵 (워크시트 자동주입용; coverage 상위3개 한계 보완, 2026.07.11)
    import re as _re2
    _disp2=getattr(load_excel,'_disp',{})
    def _nrm(s): return _re2.sub(r'[\s·()\[\]/]','',str(s))
    _dm={}
    for _rows in grp_rows.values():
        for _b,_v in _rows:
            if not _b: continue
            _dv=_disp2.get(_b) or _fmt(_v)
            if _dv and _dv!='0':
                _dm.setdefault(_nrm(_b), _dv)
    # 골절 합산(치아파절 포함+제외, 5대·수술 제외) = PPT 골절 한 값
    _gj=0
    for _rows in grp_rows.values():
        for _b,_v in _rows:
            if _b and '골절' in _b and '5대' not in _b and '수술' not in _b and isinstance(_v,(int,float)):
                _gj+=int(_v)
    if _gj>0: _dm['골절합산']=_fmt(_gj)
    rep['dambo']=_dm
    # ★v41 (지점장 2026.07.12) 보장나무(6p) CI 유동표시용 : 담보key → 금액. CI 계약일 때만 채운다.
    _ciamt={}
    if ci.get('status')=='ci':
        _cvz=_gv('중대한 뇌졸증'); _cam=_gv('중대한 급성심근'); _cca=_gv('중대한 암')
        # ★★★v251(지점장 지시 2026.07.26): <b>6p 담보별 보장범위에 CI 뇌출혈이 기재되지 않던 결함</b>.
        #   구 코드는 <b>'중대한 뇌졸증'만</b> 읽어서, v242로 축이 뇌출혈이 된 CI(신한·DB생명 실측)는
        #   그 행이 0이라 <b>hem·infarct 둘 다 CI 칩이 안 찍혔다</b>.
        #   → '중대한 뇌출혈'도 읽어 <b>뇌출혈 행(hem)</b>에 기재한다.
        #   ・중대한 뇌졸증 = 출혈+경색 포괄 → hem·infarct 둘 다 (기존 유지)
        #   ・중대한 뇌출혈 = 출혈 전용   → hem 만 (뇌경색은 미보장이므로 infarct에 넣지 않는다)
        _chm=_gv('중대한 뇌출혈')
        if _cvz: _ciamt['hem']=_fmt(_cvz); _ciamt['infarct']=_fmt(_cvz)
        if _chm: _ciamt['hem']=_fmt(_chm)
        if _cam: _ciamt['ami']=_fmt(_cam)
        if _cca: _ciamt['__cancer__']=_fmt(_cca)
    rep['ci_amounts']=_ciamt
    # ★2026.07.11 실손 세대 자동판별(CI식 3상태): 실손 계약 가입일 → 세대
    def _gen_of(js, comp='', prod='', np3=False, drug=False):
        import re as _r
        m=_r.search(r'(\d{4})\D+(\d{1,2})(?:\D+(\d{1,2}))?', str(js))
        from datetime import date
        _bycode=False
        _pm0=_r.search(r'(?<!\d)(0[9]|1[0-9]|2[0-6])\.?(0[1-9]|1[0-2])(?!\d)', str(prod or ''))
        if m:
            y=int(m.group(1)); mo=int(m.group(2)); d=int(m.group(3) or 1)
            # ★v200: main.py `silson_gen`과 동일 — 가입일과 상품코드가 둘 다 있으면 <b>더 이른 쪽</b>을
            #   쓴다(갱신 재가입일로 세대를 오판하는 것 차단). 엑셀·설명서 세대가 갈리면 안 된다.
            if _pm0:
                _y2=2000+int(_pm0.group(1)); _m2=int(_pm0.group(2))
                if (_y2,_m2,1) < (y,mo,d):
                    y,mo,d=_y2,_m2,1; _bycode=True
        else:
            # ★★v200 (윤*관 실측 2026.07.23): 가입일이 공란인 실손 계약은
            #   <b>상품명 끝의 상품코드(YYMM)</b>로 세대를 판정한다 — 지침 v90 정본을
            #   main.py `silson_gen`에만 넣고 여기(설명서·진단서 경로)에 빠뜨려
            #   '무배당 한화실손의료보험(갱신형)(실손전환용)<b>2301</b>'이 '확인불가'로 떴다.
            #   정규식·경계는 main.py `silson_gen`과 동일해야 한다(4대 산출물 연동).
            if not _pm0:
                # ★★★★★v381 (지점장 확정 2026.08.11, 영구): <b>가입일·상품코드가 둘 다 없어도
                #   담보 구조로 세대를 판정한다.</b> 지점장 원문 = "<b>4세대는 통원비 20만원만 있다</b>".
                #   → ①3대비급여(도수·주사·MRI) 특약이 분리돼 있으면 <b>3세대 이상</b>(2017.04~ 신설).
                #     ②그중 <b>처방조제료(약값)가 통원과 별도로</b> 있으면 4세대가 아니다 → <b>3세대</b>.
                #     ③통원만 있고 약값이 없으면 <b>4세대</b>.
                #   [구 결함] 날짜가 없으면 무조건 None → 그 계약은 세대 [확인]이 되고,
                #   대표 계약이 <b>다른 계약</b>으로 넘어가 세대가 통째로 오판됐다(실측 구본칠 = 1세대).
                if np3:
                    return {'gen':(3 if drug else 4),'sub':'','date':'','src':'dambo'}
                return None
            y=2000+int(_pm0.group(1)); mo=int(_pm0.group(2)); d=1; _bycode=True
        try: dt=date(y,mo,d)
        except Exception: return None

        # ★2026.07.11 지점장 확정: 1세대=~2009.09 / 2세대=2009.10~ (2009.07~09은 회사별 상이 → 주석)
        #   2세대 3분할(2-1 ~2012.12 / 2-2 2013.01~2015.12 / 2-3 2016.01~2017.03)
        #   1세대는 생보·손보 구분(자기부담 손보0%·생보20%, 상해의료비 별도)
        _LIFE=('생명','AIA','ABL','푸본','교보','동양','미래에셋','신한','KDB','메트라이프','처브','라이나','KB라이프','라이프플래닛','하나생명','IBK연금')
        _NONLIFE=('손보','손해','화재','해상','손해보험')
        sub=''
        if dt<=date(2009,9,30):
            g=1
            _c1=str(comp).split('\n')[0].strip()   # ★회사명 줄만(상품명의 '라이프' 오매칭 차단)
            if any(k in _c1 for k in _NONLIFE): sub='손보'
            elif any(k in _c1 for k in _LIFE): sub='생보'
            else: sub='손보'
        elif dt<=date(2017,3,31):
            g=2
            if dt<=date(2012,12,31): sub='2-1'
            elif dt<=date(2015,12,31): sub='2-2'
            else: sub='2-3'
        elif dt<=date(2021,6,30): g=3
        elif dt<=date(2026,5,5): g=4       # ★v211 5세대 출시=2026.05.06(금융위·금감원 보도자료 2026.5.6) — main.py silson_gen과 통일
        else: g=5
        # ★★★v250 3대비급여 하한(지점장 확정 2026.07.26) — main.py `silson_gen`과 동일 규칙.
        #   상품코드가 실손이 아닌 <b>주계약 코드</b>일 때의 세대 오판을 막는다(DB생명 CI종신1701 실측:
        #   코드 1701=2017.01로 2세대 오판 → 3대비급여 특약 분리이므로 3세대).
        #   ★날짜는 <b>원본 가입일을 그대로 표시</b>한다 — 하한 날짜(2017.04.01)를 찍으면 고객 문서에
        #     실제와 다른 가입일이 나간다(실측으로 발견).
        if np3 and g < 3:
            g=3; sub=''
        dstr=(f'{y}.{mo:02d}' if _bycode else (f'{y}.{mo:02d}.{d:02d}' if m.group(3) else f'{y}.{mo:02d}'))
        return {'gen':g,'sub':sub,'date':dstr,'src':('code' if _bycode else 'join')}
    _sil=[h for h in headers if h.get('sil')]
    if not _sil:
        rep['silson_gen']={'status':'none'}
    else:
        # ★v89 수정(장혜경 실데이터): 가입일이 빈 계약이 있으면 ''가 최솟값이 되어
        #   그 계약이 '가장 오래된 계약'으로 뽑혀 세대 판별이 통째로 실패했다.
        #   → 가입일이 있는 계약 중에서 고르고, 전부 비었을 때만 첫 계약을 쓴다.
        _dated=[h for h in _sil if str(h.get('join','')).strip()]
        if not _dated:   # ★v200 가입일이 전부 비면 상품코드로 세대가 나오는 계약을 대표로
            _dated=[h for h in _sil if _gen_of('', h.get('nm',''), h.get('prod',''), h.get('np3',False), h.get('drug',False))]
        _sh=min(_dated, key=lambda h:str(h.get('join',''))) if _dated else _sil[0]
        _cnm=_sh.get('nm','')
        _g=_gen_of(_sh.get('join'), _cnm, _sh.get('prod',''), _sh.get('np3',False), _sh.get('drug',False))
        # 실손 계약 전체 목록(회사·상품명·가입일·보험료)
        # ★v79 실손은 2개 이상일 수 있다(지점장 확정 2026.07.18).
        #   예) 상해의료비 가입 후 실손을 추가로 드는 경우 / DB손보 2006년형 특수 실손 등.
        #   → 계약을 합치지 말고 <b>각각</b> 표기하고, 세대도 <b>계약별로 각각</b> 판정한다.
        _sillist=[]
        for _h in sorted(_sil, key=lambda x:str(x.get('join',''))):
            _gh=_gen_of(_h.get('join'), _h.get('nm',''), _h.get('prod',''), _h.get('np3',False), _h.get('drug',False))
            if not _gh and _h.get('genhint'):
                _gh={'gen':_h['genhint'],'sub':'','date':''}   # ★가입일이 비면 헤더 표기 세대 사용
            _sillist.append({'co':str(_h.get('co',''))[:14],
                             'prod':str(_h.get('prod',''))[:34],
                             'join':str(_h.get('join','')),
                             'amt':_h.get('amt',0),
                             'renew':_h.get('renew',False),
                             'gen':(_gh['gen'] if _gh else ''),
                             'sub':(_gh['sub'] if _gh else '')})
        rep['silson_count']=len(_sillist)
        rep['silson_list']=_sillist
        if _g:
            rep['silson_gen']={'status':'auto','gen':_g['gen'],'sub':_g['sub'],'date':_g['date'],
                               'company':str(_sh.get('co',''))[:14],
                               'product':str(_sh.get('prod',''))[:34]}
        else:
            rep['silson_gen']={'status':'check','company':_cnm[:12]}
    return rep

if __name__=='__main__':
    import sys
    rep=map_excel_to_report(sys.argv[1] if len(sys.argv)>1 else '보장진단_정기철.xlsx',
        settings={'client':'정기철','branch':'온빛센터 바름지점','manager':'최은혜','title':'지점장','phone':'010-XXXX-XXXX'})
    for d in rep['donuts']: print(f"  {d['name']:<8} {d['pct']}%")
    print('계약',rep['n_contract'],'/ 보험료',f"{rep['premium']:,}",'/ 공백',rep['gap_count'])
