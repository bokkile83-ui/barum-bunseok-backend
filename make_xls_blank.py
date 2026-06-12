# -*- coding: utf-8 -*-
import re
from openpyxl import load_workbook
SRC='26_02월_보장분석.xlsx'; OUT='MASTER_보장분석_엑셀_영구표본.xlsx'
wb=load_workbook(SRC)
form='석준영님'
# 1) 정본 시트만 남김
for sn in list(wb.sheetnames):
    if sn!=form: wb.remove(wb[sn])
ws=wb[form]; ws.title='보장분석'

last=ws.max_column           # 합계 열
data_cols=range(3,last)      # 계약 열 C..(합계-1)
SLASH={56,62}                # 1~5종 행(상해56·질병62)

cleared=0
# 2) 이름/계약헤더 비움 (A1=이름, 1행 계약헤더, 합계 라벨은 유지)
ws.cell(1,1).value=None
for c in data_cols:
    if ws.cell(1,c).value not in (None,''): ws.cell(1,c).value=None; cleared+=1
# 3) 모든 계약 값 비움 (A구분·B담보·합계열 SUM은 보존)
for r in range(2, ws.max_row+1):
    for c in data_cols:
        if ws.cell(r,c).value not in (None,''): ws.cell(r,c).value=None; cleared+=1
    # 1~5종 합계열 = 슬래시 골격 유지
    if r in SLASH:
        ws.cell(r,last).value='/ / / / /'
wb.save(OUT)
print('SAVED',OUT,'| 비운 셀',cleared,'| 합계열',last,'| 계약열 C~',chr(64+last-1))
# 검증: 구분·담보·합계공식 보존 확인
wb2=load_workbook(OUT)
w=wb2['보장분석']
print('구분 보존:', [w.cell(r,1).value for r in (6,11,15,28,35,54,69,75,84,87)])
print('담보 보존(앞6):', [w.cell(r,2).value for r in range(6,12)])
print('합계공식 J6:', w.cell(6,last).value, '| 1~5종 J56:', w.cell(56,last).value)
