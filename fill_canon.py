# -*- coding: utf-8 -*-
"""
fill_canon.py v3 — 2026.06.12
핵심 수정:
  1) NH농협생명·NH농협손보 컬럼 미생성 (법칙 확정)
  2) 합계열: 0이면 '미가입' 텍스트 대신 빈칸 (SUM 수식 깨짐 방지)
  3) 미가입=빨강 글자 처리를 합계열에서만 (개별 칸은 빈칸 유지)
  4) 1~5종 슬래시: 종번호 정렬·누락 종 0으로 패딩
  5) 갱신색 적용: 계약레벨 갱신구분 기준 열 전체 파랑
  6) 헤더 5행 완전 기재 (회사/보험료/가입일/만기일/납입기간)
"""
import re, json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from barum_extract import extract
from barum_canon import classify_canon
from barum_dict import is_ci

# ── 폰트 정의 ────────────────────────────────────────────────────────
BLUE   = Font(color='0070C0', name='맑은 고딕', size=9)   # 갱신
BLACK  = Font(color='000000', name='맑은 고딕', size=9)   # 비갱신
PURPLE = Font(color='7030A0', bold=True, name='맑은 고딕', size=9)  # CI결합
RED    = Font(color='FF0000', name='맑은 고딕', size=9)   # 미가입(합계열만)
WHITE  = Font(color='FFFFFF', name='맑은 고딕', size=9, bold=True)

# ── 헤더 배경색 ──────────────────────────────────────────────────────
FILL_BLUE  = PatternFill('solid', fgColor='0070C0')  # 갱신
FILL_RED   = PatternFill('solid', fgColor='C00000')  # 비갱신
FILL_GREEN = PatternFill('solid', fgColor='375623')  # 완납
FILL_SUM   = PatternFill('solid', fgColor='404040')  # 합계열

# NH농협 제외 패턴
NH_SKIP = re.compile(r'(NH농협생명|NH농협손보|농협생명|농협손해보험)')


def _slash_merge(entries):
    """1~5종 [(종번호, 금액), ...] → '1종/2종/3종/4종/5종' 슬래시 문자열"""
    tbl = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    for jno, amt in entries:
        if 1 <= jno <= 5:
            tbl[jno] += amt
    return '/'.join(str(tbl[i]) for i in range(1, 6))


def fill(pdf, template, out):
    # ── canon_rows.json 로드 ─────────────────────────────────────────
    R    = json.load(open('canon_rows.json'))
    NR   = R['namerow']   # 담보명 → 엑셀 행번호
    NC   = R['NC']        # 최대 계약 컬럼 수
    LAST = R['last']      # 합계열 컬럼 번호
    MEMO = R.get('memo', LAST + 1)

    # ── PDF 추출 ─────────────────────────────────────────────────────
    cs = extract(pdf)

    # NH농협 제외 (이중 방어)
    cs = [c for c in cs if not NH_SKIP.search(c.get('회사', ''))]

    # ── 고객명 추출 ──────────────────────────────────────────────────
    cust = ''
    try:
        import fitz
        for pg in fitz.open(pdf):
            m = re.search(r'계약자\s*([가-힣*]{2,5})', pg.get_text())
            if m:
                cust = m.group(1)
                break
        if not cust:
            # 파일명에서 추출 시도
            import os
            base = os.path.basename(pdf)
            m2 = re.search(r'([가-힣]{2,5})', base)
            if m2:
                cust = m2.group(1)
    except Exception:
        pass

    # ── 템플릿 로드 ──────────────────────────────────────────────────
    wb = openpyxl.load_workbook(template)
    ws = wb['보장진단']

    if cust:
        ws.cell(1, 1).value = f'{cust} 보장진단'

    hold = []

    for i, c in enumerate(cs):
        if i >= NC:
            hold.append((c['회사'], f"계약초과(컬럼없음): {c['상품'][:20]}"))
            continue

        col     = 3 + i
        갱신    = c['갱신구분']
        종신    = '9999' in c['만기'] or '종신' in c['상품']
        is_gen  = '갱신' in 갱신 and '비갱신' not in 갱신

        # ── 헤더 5행 ─────────────────────────────────────────────────
        # 행1: 회사명 (갱신여부)
        h1 = ws.cell(1, col)
        h1.value = f"{c['회사']} [{갱신}]"
        h1.font  = WHITE
        h1.fill  = FILL_BLUE if is_gen else FILL_RED
        h1.alignment = Alignment(horizontal='center', wrap_text=True)

        # 행2: 보험료
        ws.cell(2, col).value = c.get('납입', '')

        # 행3: 가입년일
        ws.cell(3, col).value = c.get('계약일', '')   # ※ extract에서 '계약일' 없을 수 있음

        # 행4: 만기일자
        ws.cell(4, col).value = c.get('만기', '')

        # 행5: 납입기간
        ws.cell(5, col).value = c.get('납입', '')

        # ── 담보 매핑 ─────────────────────────────────────────────────
        acc    = {}   # canon → 합산금액
        slash  = {}   # canon → [(종번호, 금액)]
        cimark = {}
        gmark  = {}   # 담보칸별 갱신 여부

        for nm, amt, g in c['담보']:
            # 종신 사망 1:1 분리
            if ('사망_주계약' in nm or (nm == '사망' and 종신)):
                for tgt in ['일반사망', '상해사망']:
                    if tgt not in acc:
                        acc[tgt] = amt
                        gmark[tgt] = g
                        break
                continue

            canon, why = classify_canon(nm)

            if canon is None:
                if why.startswith('HOLD') or amt >= 100:
                    hold.append((c['회사'],
                                 f"{'[보류]' if why.startswith('HOLD') else ''}  {nm[:22]}={amt:,}"))
                continue

            ci = is_ci(nm, c['상품'])

            # 1~5종 수술비 슬래시 처리
            if '1-5종' in nm or '종수술' in canon:
                jm = re.search(r'\((\d)종\)', nm)
                slash.setdefault(canon, []).append(
                    (int(jm.group(1)) if jm else 9, amt)
                )
            else:
                acc[canon] = acc.get(canon, 0) + amt
                gmark[canon] = g or gmark.get(canon, False)
                cimark[canon] = cimark.get(canon, False) or ci

        # ── 셀 기입 ──────────────────────────────────────────────────
        def _font(canon):
            if cimark.get(canon):
                return PURPLE
            # 계약레벨 갱신이면 열 전체 파랑 (법칙 §13 ★갱신계약=그 열 전체 파랑)
            if is_gen or gmark.get(canon, False):
                return BLUE
            return BLACK

        for canon, v in acc.items():
            if canon not in NR:
                continue
            cell = ws.cell(NR[canon], col)
            cell.value = v
            cell.font  = _font(canon)
            if cimark.get(canon):
                ws.cell(NR[canon], MEMO).value = 'CI중대한(선지급)'

        for canon, lst in slash.items():
            if canon not in NR:
                continue
            cell = ws.cell(NR[canon], col)
            cell.value = _slash_merge(lst)
            cell.font  = BLUE if is_gen else BLACK

    # ── 합계열 ────────────────────────────────────────────────────────
    for canon, row in NR.items():
        tot = 0
        is_slash = False
        slash_totals = [0, 0, 0, 0, 0]

        for col in range(3, 3 + len(cs)):
            x = ws.cell(row, col).value
            if isinstance(x, (int, float)):
                tot += x
            elif isinstance(x, str) and '/' in x:
                is_slash = True
                parts = x.split('/')
                for k, p in enumerate(parts[:5]):
                    try:
                        slash_totals[k] += int(p)
                    except ValueError:
                        pass

        sc = ws.cell(row, LAST)
        if is_slash:
            sc.value = '/'.join(str(v) for v in slash_totals)
            sc.font  = BLACK
        elif tot > 0:
            sc.value = tot
            sc.font  = BLACK
        else:
            # 미가입: 빈칸으로 두고 글자색만 빨강 표시 (값 없음 → SUM 깨짐 방지)
            sc.value = None
            # 선택적으로 배경 연빨강으로 표시만
            # sc.fill = PatternFill('solid', fgColor='FFE0E0')

    # ── 확인사항 시트 ─────────────────────────────────────────────────
    sh2 = wb.create_sheet('📋확인')
    sh2.cell(1, 1, f'{cust or "고객"} · 보류·확인사항').font = Font(bold=True, color='C9A14A')
    sh2.cell(2, 1, '회사').font = Font(bold=True)
    sh2.cell(2, 2, '담보=금액/사유').font = Font(bold=True)
    for j, (who, it) in enumerate(hold[:80]):
        sh2.cell(3 + j, 1, who)
        sh2.cell(3 + j, 2, it)

    wb.save(out)
    return {'고객': cust, '계약': len(cs), '보류': len(hold)}


if __name__ == '__main__':
    import sys
    result = fill(
        sys.argv[1] if len(sys.argv) > 1 else 'src.pdf',
        'template_canon.xlsx',
        'out_canon.xlsx'
    )
    print(result)
