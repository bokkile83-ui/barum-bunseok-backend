# -*- coding: utf-8 -*-
"""
engine_v2 — 바름 보장분석 재현 엔진
  · 법칙은 전부 rules.py(정본)에서 가져온다. 이 파일엔 '고객 데이터'와 '조립'만.
  · 버그2 봉합: setc(시트, ...) — 시트를 인자로 받아 헤더 오염 제거.
  · 버그3 봉합: 라벨을 10~99에서 읽고 그대로 10~99에 쓴다(멱등). 재실행해도 안 밀림.
실행:  python3 engine_v2.py   →  보장진단_두호.xlsx 재생성
"""
import fitz, re
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
import rules

PDF = 'noname__7_.pdf'
TEMPLATE = '보장진단_두호.xlsx'   # 라벨 소스(깨끗한 표준 폼). 멱등이라 자기출력도 가능.
doc = fitz.open(PDF)

# ===== 고객 데이터 (두호) — 법칙 아님. 추후 PDF 자동추출로 대체 가능 =====
PAGES = {'C':[9],'D':[10],'E':[11],'F':[12],'G':[13],'H':[14],'I':[15,16],
         'J':[17],'K':[18,19],'L':[20,21,22,23],'M':[25,26],'N':[27],'O':[28],'P':[29,30]}
# (회사, 상품명, 가입일, 만기일, 납입년, 납입횟수, 월보험료)
META = {
 'C':('라이나생명','무배당THE간편한건강보험(갱신형)','2023.11.13','2033.11.13',10,'31/120',29240),
 'D':('동양생명','(무)수호천사플러스상해','2014.10.21','2075.10.21',20,'140/240',28300),
 'E':('동양생명','(무)수호천사내가만드는보장보험(순수보장형-종신)','2021.04.19','9999.12.31',30,'62/360',62350),
 'F':('동양생명','(무)수호천사NEW알뜰플러스종신보험','2024.05.29','9999.12.31',10,'25/120',516000),
 'G':('동양생명','(무)수호천사NEW알뜰플러스종신보험','2025.12.24','9999.12.31',7,'6/84',317500),
 'H':('한화손보','무배당 한아름플러스 종합보험(GA)','2010.03.30','2076.03.30',20,'195/240',36500),
 'I':('롯데손해보험','무배당 let:smile 4060종합보험(2604)','2026.05.27','2066.05.27',20,'1/240',125318),
 'J':('삼성화재','무배당 삼성화재 건강보험 새시대건강파트너(1004.2)','2010.08.31','2076.08.31',46,'190/552',93310),
 'K':('삼성화재','무배당 삼성 올라이프 Super보험Ⅲ(s0904)','2009.07.29','2076.07.29',67,'203/804',184730),
 'L':('KB손보','KB 5.10.10 플러스 건강보험(세만기)(무배당)(26.01)','2026.05.27','2066.05.27',20,'1/240',157801),
 'M':('DB손보','참좋은운전자상해보험2510','2025.12.08','2045.12.08',20,'6/240',35000),
 'N':('라이나(에이스)손해보험','(무)Chubb 초간편 3대질병보장보험(갱신형) 2종','2021.07.14','2041.07.14',20,'59/240',28390),
 'O':('라이나(에이스)손해보험','(무)Chubb 초간편 3대질병보장보험(갱신형) 2종','2023.09.05','2043.09.05',20,'33/240',23710),
 'P':('라이나(에이스)손해보험','(무)Chubb 치아안심보험1904 (R형/만기환급금50만원)','2020.08.18','2030.08.18',10,'70/120',36370),
}
COLS = list('CDEFGHIJKLMNOP')
고객명 = '두*호'

# ===== 1) PDF 직독 + dedup (법칙 4·dedup) =====
def pairs(col):
    skip = ('정상계약','계약자','납입주기','보험료','보장기간','월납','GA4지점',
            '최은혜','2026.06.03','/31','인천')
    out = []
    for pidx in PAGES[col]:
        ls = [l for l in doc[pidx].get_text('text').split('\n') if l.strip()
              and not any(s in l for s in skip) and ('~' not in l or '종' in l)][2:]
        name = ''
        for t in ls:
            t = t.strip(); amt = t.replace(',', '')
            if re.fullmatch(r'-?\d+', amt):
                if name: out.append((name.strip(), int(amt))); name = ''
            else:
                name = (name + ' ' + t) if name else t
    seen = set(); ded = []
    for n, a in out:
        if (n, a) in seen: continue
        seen.add((n, a)); ded.append((n, a))
    return ded

DATA = {c: pairs(c) for c in PAGES}

# ===== 2) 갱신판정 (법칙 12조, rules.판정) =====
STAT = {}
for c in COLS:
    회, 상, d1, d2, 납, 횟, _ = META[c]
    가입년 = d1[:4]; 만기년 = d2[:4]; 총회차 = int(횟.split('/')[1])
    STAT[c] = rules.판정(상, 가입년, 만기년, 납, 총회차)

# ===== 3) 담보 → 표준칸 적재 (법칙 7·9·10) =====
#   cell[행][열] = [비갱신합, 갱신합]  (13조 칸>행: 같은 칸 갱신/비갱신 분리)
cell = defaultdict(lambda: defaultdict(lambda: [0, 0]))

# ===== 동적 행매핑 (폼 행추가 안전: 구폼 담보명 → 신폼 행) =====
import os as _os
_nowwb = load_workbook(TEMPLATE)['고객 보장진단']
NOW = {str(_nowwb.cell(r, 2).value).strip(): r
       for r in range(1, _nowwb.max_row + 1) if _nowwb.cell(r, 2).value}
FORM_MAXROW = _nowwb.max_row
_BAK = '/home/claude/work/orig/보장진단_두호.bak.xlsx'
if _os.path.exists(_BAK):
    _bw = load_workbook(_BAK)['고객 보장진단']
    _BASE = {r: str(_bw.cell(r, 2).value).strip()
             for r in range(1, _bw.max_row + 1) if _bw.cell(r, 2).value}
    REMAP = {r: NOW[l] for r, l in _BASE.items() if l in NOW}   # 구행 → 신행
else:
    REMAP = {}
종수술 = defaultdict(dict); 종수술g = defaultdict(bool)
hold = {'전이': [], '116': []}
unmapped = defaultdict(list)

def 적재(r, col, amt, g):
    is_max = (isinstance(r, int) and r in rules.MAX_ROWS) or (isinstance(r, str) and r in getattr(rules,'MAX_NAMES',set()))
    if isinstance(r, str): r = NOW.get(r)
    elif isinstance(r, int): r = REMAP.get(r, r)
    if r is None: return
    i = 1 if g else 0
    if is_max:                              # ③ 합산 금지 → 대표 1건(최댓값)
        cell[r][col][i] = max(cell[r][col][i], amt)
    else:
        cell[r][col][i] += amt

for col in DATA:
    종done = False
    for name, amt in DATA[col]:
        r = rules.maprow(name); g = rules.갱신마커(name) or rules.갱신강제(name) or (STAT[col][0] == '갱신')
        if col in rules.라이나_뇌혈관_뇌출혈 and '뇌혈관' in name and '수술' not in name and '진단' in name:
            r = '뇌출혈진단비'   # 법칙29: 라이나 뇌혈관진단 = 실제 뇌출혈(한장보장표)
        if r is None:
            unmapped[col].append((name, amt)); continue
        if isinstance(r, tuple):
            tag = r[0]
            if tag == '종신':
                if 종done: continue
                종done = True
                적재(rules.일반사망_ROW, col, amt, g)
                적재(rules.상해사망_ROW, col, amt, g)
            elif tag == '전이':
                hold['전이'].append((col, name, amt))
            elif tag == '116':
                hold['116'].append((col, name, amt))
            else:
                row = rules.종상해_ROW if tag == '종상해' else rules.종질병_ROW
                row = REMAP.get(row, row)
                m = re.search(r'[_(](\d)종', name)
                종번 = m.group(1) if m else '?'
                종수술[(row, col)][종번] = amt
                종수술g[(row, col)] |= g
            continue
        적재(r, col, amt, g)

def 칸합(r, cc):
    return cell[r][cc][0] + cell[r][cc][1]

# 법칙23 — 유사암 미기재 회사 자동 추정 (① 기재면 직독 유지 / ② 미기재면 가입연도 기준 ÷10·÷5)
유사암추정 = {}
_유사행 = REMAP.get(rules.유사암_검산행, rules.유사암_검산행)
_일반행 = REMAP.get(rules.일반암_행, rules.일반암_행)
_일반갱 = NOW.get('일반암(갱신형)')
for cc in COLS:
    if 칸합(_유사행, cc) == 0:          # 미기재
        일반암 = 칸합(_일반행, cc) + (칸합(_일반갱, cc) if _일반갱 else 0)
        if 일반암:
            est = rules.유사암추정(일반암, META[cc][2][:4])
            cell[_유사행][cc] = [est, 0]  # 추정값 기재(검정)
            유사암추정[cc] = (est, META[cc][2][:4])

# ===== 4) 라벨 읽기 (멱등: 10~99에서 읽어 10~99에 쓴다) =====
tpl = load_workbook(TEMPLATE)['고객 보장진단']
labels = {r: (tpl[f'A{r}'].value, tpl[f'B{r}'].value)
          for r in range(rules.DATA_START, FORM_MAXROW + 1)}

# ===== 5) 폼 조립 =====
wb = Workbook(); ws = wb.active; ws.title = '고객 보장진단'
thin = Side(style='thin', color='BBBBBB'); BORDER = Border(thin, thin, thin, thin)

def setc(sheet, ref, val, font=None, fill=None, align='center'):   # ★버그2 봉합: 시트 인자
    c = sheet[ref]; c.value = val
    if font: c.font = font
    if fill: c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = BORDER

F = lambda **k: Font(name='맑은 고딕', **k)
HF = F(color=rules.HEADER_FONT_COLOR, bold=True, size=9)
LBL = F(bold=True, size=9)
COL = {'갱신': F(color=rules.VALUE_COLOR['갱신'], size=9),
       '비갱신': F(color=rules.VALUE_COLOR['비갱신'], size=9)}
RED = F(color='FF0000', size=9); GREEN = F(color='008000', bold=True, size=9)

# 헤더 좌측 라벨 (B1~B8)
for i, l in enumerate(rules.LEFT_LABELS, 1):
    setc(ws, f'B{i}', l, LBL, PatternFill('solid', fgColor='D6DCE5'))
setc(ws, 'A1', f'{고객명} 고객님', F(bold=True, size=11))
setc(ws, 'Q1', '합계', HF, PatternFill('solid', fgColor='1F3864'))
setc(ws, 'R1', '4P검산', HF, PatternFill('solid', fgColor='1F3864'))

# 계약 헤더 (C~P)
for cc in COLS:
    회, 상, d1, d2, 납, 횟, 보험료 = META[cc]
    st, 근거 = STAT[cc]
    fill = PatternFill('solid', fgColor=rules.HEADER_FILL.get(st, 'C00000'))
    setc(ws, f'{cc}1', 회, HF, fill)
    setc(ws, f'{cc}2', 상, F(size=8))
    setc(ws, f'{cc}3', d1, COL['비갱신'])
    setc(ws, f'{cc}4', d2 if d2[:4] != '9999' else '종신(9999)', COL['비갱신'])
    setc(ws, f'{cc}5', f'월납/{납}년', COL['비갱신'])
    setc(ws, f'{cc}6', 횟, COL['비갱신'])
    setc(ws, f'{cc}7', 보험료, COL['비갱신']); ws[f'{cc}7'].number_format = '#,##0'
    setc(ws, f'{cc}8', st, COL['갱신'] if st == '갱신' else RED)
setc(ws, 'Q7', '=SUM(C7:P7)', COL['비갱신']); ws['Q7'].number_format = '#,##0'
ws.row_dimensions[2].height = 28; ws.row_dimensions[9].height = 6

# 값 셀 출력 (① 갱신/비갱신 슬래시 색분리: 비갱신 검정 / 갱신 파랑)
GFONT = InlineFont(rFont='맑은 고딕', color='0000FF', sz=9)
BFONT = InlineFont(rFont='맑은 고딕', color='000000', sz=9)
def put_value(ref, bv, gv):
    if not bv and not gv: return
    c = ws[ref]
    if bv and gv:
        c.value = CellRichText([TextBlock(BFONT, str(bv)), TextBlock(BFONT, '/'), TextBlock(GFONT, str(gv))])
    elif gv:
        c.value = gv; c.font = COL['갱신']
    else:
        c.value = bv; c.font = COL['비갱신']
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER

# 담보 90행+ (멱등, 폼 행추가 자동수용)
P4N = {REMAP.get(k, k): (lab, tgt, [REMAP.get(x, x) for x in rowsum])
       for k, (lab, tgt, rowsum) in rules.P4.items()}
미가입N = {REMAP.get(x, x) for x in rules.미가입행}
_ig = NOW.get('일반암(갱신형)')                       # 일반암 검산 = 22+23(갱신형) 합
if _ig:
    for _k in list(P4N):
        _lab,_tgt,_rows = P4N[_k]
        if REMAP.get(22,22) in _rows: P4N[_k]=(_lab,_tgt,_rows+[_ig])
유사검N = REMAP.get(rules.유사암_검산행, rules.유사암_검산행)
for r in range(rules.DATA_START, FORM_MAXROW + 1):
    cat, nm = labels[r]
    if cat: setc(ws, f'A{r}', cat, LBL, PatternFill('solid', fgColor='EDEDED'))
    setc(ws, f'B{r}', nm, F(size=9), align='left')
    # 종수술 행(슬래시) 우선
    if any(rk == r for (rk, _) in 종수술):
        tot = [0] * 5
        for (rk, cc), d in 종수술.items():
            if rk == r:
                s = '/'.join(str(d.get(str(k), 0)) for k in range(1, 6))
                setc(ws, f'{cc}{r}', s, COL['갱신'] if 종수술g[(rk, cc)] else COL['비갱신'])
                for k in range(1, 6): tot[k - 1] += d.get(str(k), 0)
        setc(ws, f'Q{r}', '/'.join(map(str, tot)), COL['비갱신'])
    else:
        for cc in COLS:
            put_value(f'{cc}{r}', cell[r][cc][0], cell[r][cc][1])
        tot = sum(칸합(r, cc) for cc in COLS)
        setc(ws, f'Q{r}', tot if tot else None, COL['비갱신'])
    # R 검산
    if r in P4N:
        lab, tgt, rowsum = P4N[r]
        s = sum(칸합(rr, cc) for rr in rowsum for cc in COLS)
        s += sum(sum(d.values()) for (rk, _), d in 종수술.items() if rk in rowsum)
        if r == 유사검N:        # 법칙23: 유사암은 기재값+추정으로 이미 채움
            s = sum(칸합(r, cc) for cc in COLS)
            est = ','.join(f'{cc}÷{10 if int(y)<2021 else 5}={v}' for cc,(v,y) in 유사암추정.items())
            setc(ws, f'R{r}', f'법칙23 합 {s}' + (f' (추정: {est})' if est else ' (전부 기재)'), GREEN, align='left')
            continue
        if s == tgt: setc(ws, f'R{r}', f'✅ {lab} {tgt}', GREEN, align='left')
        elif s > tgt: setc(ws, f'R{r}', f'⚠초과 {s}/{lab} {tgt} [확인]', RED, align='left')
        else: setc(ws, f'R{r}', f'⚠미달 {s}/{lab} {tgt} [확인]', RED, align='left')
    elif r in 미가입N:
        setc(ws, f'R{r}', '미가입(4P)', RED, align='left')

# 116대 단계별 슬래시 (Ⅰ~Ⅵ = 1-5종보다 상위 등급)
_116row = NOW.get('116대수술비')
_rmn = {'Ⅰ': 1, 'Ⅱ': 2, 'Ⅲ': 3, 'Ⅳ': 4, 'Ⅴ': 5, 'Ⅵ': 6}
_116d = defaultdict(dict)
for col, name, amt in hold['116']:
    gnum = next((v for k, v in _rmn.items() if k in name), None)
    if gnum: _116d[col][gnum] = amt
if _116row:
    for cc in COLS:
        d = _116d.get(cc)
        if d:
            s = '/'.join(str(d.get(k, 0)) for k in range(1, 7))
            setc(ws, f'{cc}{_116row}', s, COL['갱신'] if STAT[cc][0] == '갱신' else COL['비갱신'])

# 너비 / 고정
ws.column_dimensions['A'].width = 8; ws.column_dimensions['B'].width = 20
for cc in COLS: ws.column_dimensions[cc].width = 10
ws.column_dimensions['Q'].width = 11; ws.column_dimensions['R'].width = 26
ws.freeze_panes = 'C10'

# ===== 6) 📋확인 시트 (18조: 자수) — ★setc에 chk 전달 =====
chk = wb.create_sheet('📋확인(규칙필요)')
for c, h in zip('ABC', ['행/항목', '상태', '후보·근거 — 규칙 1줄씩']):
    setc(chk, f'{c}1', h, F(bold=True), PatternFill('solid', fgColor='D6DCE5'), 'left')
rr = 2
def line(a='', b='', c='', bold=False):
    global rr
    chk[f'A{rr}'] = a; chk[f'B{rr}'] = b; chk[f'C{rr}'] = c
    if bold: chk[f'A{rr}'].font = F(bold=True)
    rr += 1
line('■ 갱신판정 (법칙12조)', bold=True)
for cc in COLS:
    st, 근거 = STAT[cc]; line(f'{cc} {META[cc][0]}', st, 근거)
line()
line('■ 4P 미닫힘 = 약관 규칙 필요 (18조: 비우고 [확인])', bold=True)
for r, (lab, tgt, rowsum) in sorted(rules.P4.items()):
    s = sum(칸합(rr, cc) for rr in rowsum for cc in COLS)
    s += sum(sum(d.values()) for (rk, _), d in 종수술.items() if rk in rowsum)
    if s == tgt: continue
    line(f'행{r} {lab}', f'{"초과" if s>tgt else "미달"} {s}/{tgt}',
         ', '.join(f'{c}={칸합(rr,c)}' for rr in rowsum for c in COLS if 칸합(rr, c)))
line()
line('■ 전이암 (법칙10 대표1행) — 폼 행위치 [확인]', bold=True)
for col, name, amt in hold['전이']: line(col, amt, name)
line()
line('■ 116대 질병수술 (종수술 후보)', bold=True)
for col, name, amt in hold['116']: line(col, amt, name)
chk.column_dimensions['A'].width = 22; chk.column_dimensions['B'].width = 14
chk.column_dimensions['C'].width = 95

wb.save(TEMPLATE)
print('engine_v2 재작성 완료')
print('갱신판정:', {c: STAT[c][0] for c in COLS})
def _chk(r):
    lab, tgt, rowsum = rules.P4[r]
    s = sum(칸합(rr, cc) for rr in rowsum for cc in COLS)
    s += sum(sum(d.values()) for (rk, _), d in 종수술.items() if rk in rowsum)
    return s == tgt
ok = sum(1 for r in rules.P4 if _chk(r))
print(f'4P 검산 ✅ {ok}/{len(rules.P4)}')
