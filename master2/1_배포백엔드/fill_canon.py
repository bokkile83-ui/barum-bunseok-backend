# -*- coding: utf-8 -*-
"""캐논 template 채우기 + 색상(법칙13)·CI표식. fill(pdf,template,out)."""
import re, json, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from barum_extract import extract
from barum_canon import classify_canon
from barum_dict import is_ci

BLUE=Font(color='0070C0',name='맑은 고딕',size=9)      # 갱신
BLACK=Font(color='000000',name='맑은 고딕',size=9)     # 비갱신
PURPLE=Font(color='7030A0',bold=True,name='맑은 고딕',size=9)  # CI결합
RED=Font(color='FF0000',name='맑은 고딕',size=9)       # 미가입/0

def fill(pdf, template, out):
    R=json.load(open('canon_rows.json')); NR=R['namerow']; NC=R['NC']; LAST=R['last']; MEMO=R['memo']
    cs=extract(pdf)
    cust=''
    try:
        import fitz
        for pg in fitz.open(pdf):
            m=re.search(r'계약자\s*([가-힣*]+)',pg.get_text())
            if m: cust=m.group(1); break
    except: pass
    wb=openpyxl.load_workbook(template); ws=wb['보장진단']
    if cust: ws.cell(1,1).value=f'{cust} 보장진단'
    hold=[]
    for i,c in enumerate(cs):
        if i>=NC: hold.append((f"{c['회사']} {c['상품'][:14]}","계약초과(컬럼없음)")); continue
        col=3+i; 종신='9999' in c['만기'] or '종신' in c['상품']
        ws.cell(1,col).value=c['회사']; ws.cell(2,col).value=c['갱신구분']
        ws.cell(4,col).value=c['만기']
        acc={}; slash={}; cimark={}; gmark={}
        for nm,amt,g in c['담보']:
            if nm.startswith('사망_주계약') and 종신:
                tgt='일반사망' if '일반사망' not in acc else '상해사망'
                acc[tgt]=amt; gmark[tgt]=g; continue
            canon,why=classify_canon(nm)
            if canon is None:
                if why.startswith('HOLD') or amt>=100:
                    hold.append((c['회사'],f"{'[보류]' if why.startswith('HOLD') else ''}{nm.split('_')[0][:22]}={amt}"))
                continue
            ci=is_ci(nm,c['상품'])
            if '1-5종' in nm or '종수술' in canon:
                jm=re.search(r'\((\d)종\)',nm); slash.setdefault(canon,[]).append((int(jm.group(1)) if jm else 9,amt))
            else:
                acc[canon]=acc.get(canon,0)+amt
            gmark[canon]=g; cimark[canon]=cimark.get(canon,False) or ci
        for canon,v in acc.items():
            if canon not in NR: continue
            cell=ws.cell(NR[canon],col); cell.value=v
            cell.font = PURPLE if cimark.get(canon) else (BLUE if gmark.get(canon) else BLACK)
            if cimark.get(canon): ws.cell(NR[canon],MEMO).value='CI중대한(선지급)'
        for canon,lst in slash.items():
            if canon not in NR: continue
            cell=ws.cell(NR[canon],col); cell.value='/'.join(str(a) for _,a in sorted(lst))
            cell.font = BLUE if gmark.get(canon) else BLACK
    # 합계열
    for canon,row in NR.items():
        tot=0; has=False
        for col in range(3,3+NC):
            x=ws.cell(row,col).value
            if isinstance(x,(int,float)): tot+=x; has=True
        sc=ws.cell(row,LAST)
        if has and tot>0: sc.value=tot; sc.font=BLACK
        else: sc.value='미가입'; sc.font=RED
    # 확인사항
    sh2=wb.create_sheet('📋확인')
    sh2.cell(1,1,f'{cust or "고객"} · 보류·확인 (캐논 자동분류)').font=Font(bold=True,color='C9A14A')
    sh2.cell(2,1,'회사'); sh2.cell(2,2,'담보=금액/사유')
    for j,(who,it) in enumerate(hold[:80]): sh2.cell(3+j,1,who); sh2.cell(3+j,2,it)
    wb.save(out)
    return {'고객':cust,'계약':len(cs),'보류':len(hold)}

if __name__=='__main__':
    import sys; print(fill(sys.argv[1] if len(sys.argv)>1 else 'src.pdf','template_canon.xlsx','out_canon.xlsx'))
