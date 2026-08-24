# -*- coding: utf-8 -*-
"""제1지침 게이트 (지점장 지시 2026.08.23) — 매 턴 첫 도구 호출로 이것만 실행한다.
   ① 읽기  ② 지침=메모리 일치화 실측  ③ 날짜 박은 zip 최신본 생성
   사용:  python3 sync_check.py            (체크만)
          python3 sync_check.py --zip      (체크 + zip 생성)
   ※ 배포 14파일이 아니다. 검증 도구다."""
import os, re, sys, hashlib, subprocess, datetime

W = os.path.dirname(os.path.abspath(__file__))
OUT = '/mnt/user-data/outputs'
FILES14 = ['main.py','assets_b64.py','coverage_benchmark.py','ga_tables.py','remodel.py',
           'report_pages.py','report_pptx.py','report_weasy.py','master.xlsx','ppt_form.pptx',
           'Dockerfile','nixpacks.toml','requirements.txt','BARUM_DOCTRINE.md']

# ── 메모리 기재값 (정본). 값을 바꿀 땐 메모리도 같이 바꾼다 ──
MEM = {
    'master_md5'  : 'a963d8fa243635f0f5142828ee567ad1',
    'master_size' : 19372,
    'sheets'      : 4,
    'rows'        : 106,
    'dambo'       : 101,
    'gubun'       : 15,
    'heart'       : '협심증·심부전·빈맥·염증·부정맥·심근병증·심장판막·산정특례심장·'
                    '2대 주요치료비·허혈성 진단비·급성심근경색·중대한 급성심근·혈전용해치료비',
    'silson'      : 5,
    'ndae_row'    : 83,
    'doc_lines'   : 5958,
    'doc_md5'     : '228e09088fa36c7f8bb24e625f639a58',
    'articles'    : 131,      # 제0~130조. 제398·420조는 v398·v420 오기라 제외
    'missing'     : 0,
    'heart_cases' : 29,
    'audit_cases' : 91,      # ★v560 시트58 + _EXTRA_CASES 33
    'silson_cases': 23,
    'name_cases'  : 13,      # ★v568 이름 자가진단 케이스
    'ga_body'     : 'False',
    'ga_cover'    : 'False',
    'files14'     : 14,
    'proj_md'     : 0,
    'ji1'         : 1,       # ★DOCTRINE에 「제1지침」이 박혀 있는가
    'nogrep_noask': 1,       # ★「질문 전 grep」 조항이 박혀 있는가
    'no_feel'     : 1,       # ★「느낌을 쓰지 않는다」 0항이 박혀 있는가
    'no_master_bug': 1,      # ★「마스터 무행은 결함이 아니다」 조항
    'no_ask_excel': 1,       # ★「엑셀에 없는 건 묻지 않는다」 조항
}

def md5(p):
    return hashlib.md5(open(p, 'rb').read()).hexdigest()

def measure():
    a = {}
    stamps = set()
    for f in ['main.py', 'coverage_benchmark.py', 'report_weasy.py', 'ga_tables.py']:
        stamps |= set(re.findall(r'v\d{3}-[a-z0-9]+-\d{8}', open(os.path.join(W, f), encoding='utf-8').read()))
    a['stamp'] = sorted(stamps)

    a['master_md5']  = md5(os.path.join(W, 'master.xlsx'))
    a['master_size'] = os.path.getsize(os.path.join(W, 'master.xlsx'))
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(W, 'master.xlsx'))
    ws = wb.worksheets[0]
    a['sheets'] = len(wb.worksheets)
    a['rows']   = ws.max_row
    a['dambo']  = sum(1 for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value)
    # ★구분은 A열 값에서 헤더 4행(보험료·가입년일·만기일자·총납입기간)을 뺀다
    hdr = {'보험료', '가입년일', '만기일자', '총납입기간'}
    a['gubun'] = sum(1 for r in range(1, ws.max_row + 1)
                     if ws.cell(r, 1).value and str(ws.cell(r, 1).value).strip() not in hdr)
    a['heart']  = '·'.join(str(ws.cell(r, 2).value) for r in range(41, 54))
    a['silson'] = sum(1 for r in range(101, 106) if ws.cell(r, 2).value)
    a['ndae_row'] = 83 if ws.cell(83, 2).value else 0

    d = open(os.path.join(W, 'BARUM_DOCTRINE.md'), encoding='utf-8').read()
    a['doc_lines'] = len(d.split('\n')) - 1
    a['doc_md5']   = md5(os.path.join(W, 'BARUM_DOCTRINE.md'))
    # ★v398·v420 오기를 뺀 실조문만 센다
    arts = sorted({int(x) for x in re.findall(r'제(\d+)조', d) if int(x) <= 300})
    a['articles'] = len(arts)
    a['missing']  = len([n for n in range(arts[0], arts[-1] + 1) if n not in arts])

    src = open(os.path.join(W, 'main.py'), encoding='utf-8').read()
    ec = re.search(r'_EXTRA_CASES = \[(.*?)\n\]', src, re.S)
    a['audit_cases'] = 58 + (ec.group(1).count("),") if ec else -1)
    hc = re.search(r'_HEART_CASES\s*=\s*\[(.*?)\n\]', src, re.S)
    sc = re.search(r'_SILSON_CASES\s*=\s*\[(.*?)\n\]', src, re.S)
    a['heart_cases']  = hc.group(1).count('),') if hc else -1
    a['silson_cases'] = sc.group(1).count('),') if sc else -1
    nc = re.search(r'_NAME_CASES = \[(.*?)\n\]', src, re.S)
    a['name_cases'] = nc.group(1).count('),') if nc else -1

    g = open(os.path.join(W, 'ga_tables.py'), encoding='utf-8').read()
    a['ga_body']  = re.search(r'_GA_BODY\s*=\s*(\w+)', g).group(1)
    a['ga_cover'] = re.search(r'_GA_COVER\s*=\s*(\w+)', g).group(1)

    a['files14'] = sum(1 for f in FILES14 if os.path.exists(os.path.join(W, f)))
    a['proj_md'] = len([f for f in os.listdir('/mnt/project') if f.endswith('.md')]) \
                   if os.path.isdir('/mnt/project') else 0
    a['ji1'] = 1 if '## 제1지침 — 매 턴 무한반복' in d else 0
    a['nogrep_noask'] = 1 if '질문 금지 조항 — 지침에 있는 것은 묻지 않는다' in d else 0
    a['no_feel'] = 1 if '0항 — 느낌을 쓰지 않는다' in d else 0
    a['no_master_bug'] = 1 if '마스터 무행은 결함이 아니다' in d else 0
    a['no_ask_excel'] = 1 if '엑셀에 없는 건 묻지 않는다 — 지점장이 말하기 전까지' in d else 0
    return a

def recite():
    """★★★★★v567 (지점장 2026.08.23 「니가 아무리 검사기를 넣어도 결국 지침 안 보고 있고
       영구다 해도 안 보고 지침1번이라고 해도 안 본다」).
       [사실] `ji1`·`nogrep_noask`·`no_feel` 3항목이 전부 OK인데도 조문을 안 읽고 두 번 어겼다.
       <b>존재 확인은 읽기가 아니다.</b> ⇒ 검사기가 조문 <b>본문을 찍어낸다</b>.
       매 턴 첫 도구 호출이므로 이 글자가 매번 눈앞에 강제로 온다."""
    d = open(os.path.join(W, 'BARUM_DOCTRINE.md'), encoding='utf-8').read()
    try:
        i = d.index('## 제1지침 — 매 턴 무한반복')
        j = d.index('## 제0조 — 최상위 법률')
        body = d[i:j]
    except ValueError:
        print('★제1지침 조문을 찾지 못했다 — 즉시 멈춘다'); return 1
    print('=' * 74)
    print('제1지침 ① 낭독 — 이 글자를 읽지 않고 답하지 않는다')
    print('=' * 74)
    for ln in body.split('\n'):
        t = ln.rstrip()
        if not t or t.startswith(('|---', '> ')): continue
        print('  ' + t[:110])
    print('=' * 74)
    print('★ 이번 턴 자문 3가지 — 답변을 쓰기 전에')
    print('   1. 질문·「결함」·「확인 필요」를 쓰려는가 → grep 먼저 돌렸는가 · 출력을 붙였는가')
    print('   2. 「고쳤다·0건」을 쓰려는가 → 이 턴의 도구 출력이 붙어 있는가')
    print('   3. 산출물을 건드렸는가 → 2단(--render)까지 돌렸는가')
    return 0


def main():
    recite()
    a = measure()
    print('=' * 74)
    print('제1지침 ② 일치화 체크 —', datetime.date.today().strftime('%Y-%m-%d'))
    print('=' * 74)
    bad = 0
    for k, mem in MEM.items():
        act = a[k]
        ok = str(mem) == str(act)
        if not ok:
            bad += 1
        mark = 'OK    ' if ok else '★불일치'
        print(f'{mark} | {k:13s} | 메모리 {str(mem)[:34]:<34} | 실측 {str(act)[:34]}')
    print('-' * 74)
    print(f'각인 4파일 : {a["stamp"]}   ({"일치" if len(a["stamp"]) == 1 else "★불일치"})')
    print(f'불일치 {bad} / {len(MEM)}')

    if bad or len(a['stamp']) != 1:
        print('\n★ 불일치가 있다. zip 생성 전에 메모리 또는 코드를 맞춘다.')

    if '--zip' in sys.argv:
        if len(a['stamp']) != 1:
            print('\n각인이 갈려 zip을 만들지 않는다.'); return 1
        ver = a['stamp'][0].split('-')[0]
        today = datetime.date.today().strftime('%Y%m%d')
        name = f'MAKEONE_{ver}_{today}_FINAL.zip'
        path = os.path.join(OUT, name)
        for f in os.listdir(OUT):
            if f.endswith('.zip'):
                os.remove(os.path.join(OUT, f))
        subprocess.run(['zip', '-q', '-X', path] + FILES14, cwd=W, check=True)
        print(f'\n③ zip 생성 : {name}  {os.path.getsize(path):,}B  (14파일)')
    return 0



# ══════════════════════════════════════════════════════════════════
# 2단 게이트 — 실제 산출물 검사 (지점장 지적 2026.08.23
#   「10번이 이리 빨리 되냐」 = 위 1단은 지문만 본다. 산출물을 안 만든다)
#   1단 sync_check  0.24초  파일 5개 지문·설정
#   2단 render_check  45초  실제 45쪽을 렌더해서 쪽·제목·이미지·백지를 본다
#   ★1단만 통과한 것을 「검증됐다」고 쓰지 않는다.
# ══════════════════════════════════════════════════════════════════
PAGES = {   # pgn : 그 쪽에 반드시 있어야 할 글자
    27: '암, 보장률은 내려가고',
    28: '암 보장률',
    29: '어디서 치료받느냐',
    30: '암 치료비',
    35: '순환계 한눈에',
    36: '레켐비',
    37: '연금계좌',
}
IMG_PAGES = [30, 37]        # 그림이 실제로 박혀 있어야 하는 쪽
TOTAL, REFS = 45, 33        # 전체 쪽 · 참고자료 쪽

def render_check():
    import subprocess, glob
    print('\n' + '=' * 74)
    print('2단 게이트 — 실제 산출물 검사')
    print('=' * 74)
    for f in glob.glob(os.path.join(W, '*.pdf')):
        os.remove(f)
    r = subprocess.run(['python3', 'report_weasy.py', 'master.xlsx', 'GATE'],
                       cwd=W, capture_output=True)
    pdf = os.path.join(W, '보장설명지_GATE.pdf')
    if not os.path.exists(pdf):
        print('★렌더 실패\n' + r.stderr.decode('utf-8', 'replace')[-600:]); return 1

    n = int(subprocess.run(['pdfinfo', pdf], capture_output=True)
            .stdout.decode().split('Pages:')[1].split()[0])
    bad = 0
    def line(ok, txt):
        nonlocal bad
        if not ok: bad += 1
        print(('OK     | ' if ok else '★불량  | ') + txt)

    line(n == TOTAL, f'전체 쪽수        기준 {TOTAL}   실측 {n}')
    line(n - 12 == REFS, f'참고자료 쪽수     기준 {REFS}   실측 {n - 12}')

    def text(p):
        return subprocess.run(['pdftotext', '-layout', '-f', str(p), '-l', str(p), pdf, '-'],
                              capture_output=True).stdout.decode('utf-8', 'replace')

    for pgn, must in PAGES.items():
        t = text(pgn)
        line(must in t, f'pgn {pgn:<3d}          「{must}」 존재')

    # 그림 페이지 — 텍스트가 거의 없고 이미지가 박혀 있어야 한다
    for pgn in IMG_PAGES:
        img = subprocess.run(['pdfimages', '-list', '-f', str(pgn), '-l', str(pgn), pdf],
                             capture_output=True).stdout.decode()
        cnt = max(0, len(img.strip().split('\n')) - 2)
        line(cnt >= 1, f'pgn {pgn:<3d}          그림 {cnt}개 박힘')

    # 백지 페이지
    blank = [p for p in range(1, n + 1) if len(text(p).strip()) < 40]
    line(not blank, f'백지 페이지       기준 0     실측 {len(blank)} {blank if blank else ""}')

    os.remove(pdf)
    print('-' * 74)
    print(f'2단 불량 {bad}건')
    return bad

if __name__ == '__main__':
    rc = main()                       # 1단은 항상 돈다
    if '--render' in sys.argv:
        rc += render_check()          # 2단은 --render 일 때만
    sys.exit(rc)
