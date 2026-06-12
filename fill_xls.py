# -*- coding: utf-8 -*-
import sys, re, fitz
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import let_engine as E, meta_parser as M
MASTER='MASTER_보장분석_엑셀_영구표본.xlsx'

ROW={'질병사망':8,'상해사망':10,'질병후유3%':13,'질병후유80%':14,'상해후유3%':11,'상해후유80%':12,
 '일반암':16,'고액암':15,'암수술':25,'고액항암':19,
 '뇌혈관진단':28,'뇌졸중진단':29,'뇌출혈진단':31,'뇌혈관수술':63,'급성심근경색':41,'심장수술':65,
 '중증뇌혈관':33,'중증심장':39,'질병입원의료':84,'질병통원의료':85,'상해수술':54,'질병수술':61,
 '질병중환자실':52,'상해중환자실':53,'응급실내원':78,'화상진단':80,'골절진단':76,'깁스치료':83,
 '질병입원':44,'상해입원':47,'간병질병입원':48,'일상생활배상':87}

BLUE=PatternFill('solid',fgColor='0070C0'); RED=PatternFill('solid',fgColor='C00000')

def main(pdf):
    d=fitz.open(pdf); t,_=E.한장보장표(d); cons=M.보유계약(d)
    p0=d[0].get_text('text'); name='고객'
    for ln in [x.strip() for x in p0.split('\n') if x.strip()]:
        m=re.match(r'(.+?)\s*고객님',ln)
        if m: name=m.group(1).strip(); break
    wb=load_workbook(MASTER); ws=wb['보장분석']
    base=ws.max_column                      # 합계열(=10)
    ndata=base-3                            # 기존 계약열 수(C..합계-1 = 7)
    need=len(cons)
    if need>ndata:                          # 계약>열 → 열 추가(법칙7 추가만)
        ws.insert_cols(base, need-ndata)
    last=ws.max_column                      # 새 합계열
    # 합계 SUM 범위 갱신
    from openpyxl.utils import get_column_letter as col
    c0=col(3); cN=col(last-1)
    for r in range(2, ws.max_row+1):
        v=ws.cell(r,last).value
        if isinstance(v,str) and v.startswith('=SUM'):
            ws.cell(r,last).value=f'=SUM({c0}{r}:{cN}{r})'
    # 고객명
    ws.cell(1,1).value=f'{name} (전)'
    # 계약 헤더(제목칸) + 갱신색
    for i,cn in enumerate(cons):
        c=3+i
        h=ws.cell(1,c); h.value=f'{cn["회사"]}\n{cn["상품"][:14]}'
        h.fill=BLUE if cn['갱신']=='갱신' else RED
        h.font=Font(color='FFFFFF',bold=True,size=8); h.alignment=Alignment(wrap_text=True,vertical='center',horizontal='center')
        tcol=Font(color='0000FF' if cn['갱신']=='갱신' else '000000',size=9)
        ws.cell(2,c).value=cn['보험료']; ws.cell(2,c).font=tcol
        ws.cell(3,c).value=cn['계약일']; ws.cell(3,c).font=tcol
        ws.cell(4,c).value=cn['만기일']; ws.cell(4,c).font=tcol
        ws.cell(5,c).value=f'{cn["납입년"]}({cn["갱신"]})'; ws.cell(5,c).font=tcol
    # 담보 총액 → 합계열(계약별 분해는 별첨 이미지라 미완)
    nfill=0
    for key,r in ROW.items():
        v=t.get(key,0)
        if v: ws.cell(r,last).value=v; ws.cell(r,last).font=Font(color='000000'); nfill+=1
    note=ws.cell(1,last+2); note.value=f'[확인] 계약 {len(cons)}건 헤더·갱신색 자동(갱신=파랑/비갱신=빨강). 합계=한장보장표 담보총액. 계약별 담보 분해는 별첨 담보명 이미지라 미완(수기). 허혈성진단·1~5종·n대수술=행없음/별첨.'
    note.font=Font(color='FF0000',size=9)
    out=f'/mnt/user-data/outputs/보장진단_{name}.xlsx'
    wb.save(out); print('saved',out,f'| 계약{len(cons)} 갱신{sum(1 for c in cons if c["갱신"]=="갱신")} 담보{nfill}'); return out
if __name__=='__main__': main(sys.argv[1])
